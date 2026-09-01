import os
import csv
import sqlite3
import io
import json
from datetime import datetime, date, time, timedelta, timezone
from zoneinfo import ZoneInfo
from functools import wraps

from flask import (
    Flask, render_template, request, redirect, url_for,
    flash, jsonify, session, g, has_request_context, abort
)
from werkzeug.exceptions import HTTPException
from flask_sqlalchemy import SQLAlchemy
from flask_login import (
    LoginManager, UserMixin, login_user, login_required,
    logout_user, current_user
)
from werkzeug.security import generate_password_hash, check_password_hash
from sqlalchemy import func, and_, event
from sqlalchemy.engine import Engine
from sqlalchemy.orm import joinedload

# ── App Config ───────────────────────────────────────────────────────────────

DEV_SECRET_KEY = 'logistat-dev-key-change-me'
# Gorny limit uploadu (import CSV/Excel czyta plik do pamieci przez file.read(),
# a gunicorn ma 2 workery — bez limitu jeden plik potrafi wywalic kontener).
DEFAULT_MAX_UPLOAD_MB = 32
# Haslo konta 'admin' zakladanego przy pierwszym starcie; nadpisywalne
# zmienna ADMIN_PASSWORD (patrz docs/DEPLOY.md).
DEFAULT_ADMIN_PASSWORD = 'admin123'


def resolve_secret_key(env=None):
    """Zwroc SECRET_KEY albo rzuc, jesli tryb serwerowy leci na kluczu dev.

    Obecnosc DATABASE_URL = srodowisko test/prod (tak ustawia
    docker-compose.override.yml). Dotad brak SECRET_KEY cicho spadal na staly
    klucz dev — czyli sesje kazdego admina dalyby sie podrobic, i nikt by sie
    o tym nie dowiedzial. Awaryjnie: LOGISTAT_ALLOW_DEV_SECRET=1.
    """
    env = os.environ if env is None else env
    key = env.get('SECRET_KEY', '').strip()
    if key:
        return key
    tryb_serwerowy = bool(env.get('DATABASE_URL', '').strip())
    if tryb_serwerowy and env.get('LOGISTAT_ALLOW_DEV_SECRET') != '1':
        raise RuntimeError(
            'SECRET_KEY nie jest ustawiony, a DATABASE_URL wskazuje srodowisko '
            'serwerowe. Ustaw SECRET_KEY w docker-compose.override.yml '
            '(python3 -c "import secrets; print(secrets.token_hex(32))"). '
            'Swiadome pominiecie: LOGISTAT_ALLOW_DEV_SECRET=1.'
        )
    return DEV_SECRET_KEY


app = Flask(__name__)
app.config['SECRET_KEY'] = resolve_secret_key()
# SQLite domyslnie (dev); produkcyjnie/testowo Postgres przez DATABASE_URL,
# np. postgresql+psycopg2://logistat:...@db:5432/logistat
app.config['SQLALCHEMY_DATABASE_URI'] = os.environ.get('DATABASE_URL', 'sqlite:///logistat.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
# Sprawdz polaczenie przed uzyciem: po restarcie kontenera bazy workery trzymaja
# martwe polaczenia z puli i pierwsze zadanie na kazdym zwracalo 500
# ("server closed the connection unexpectedly"). Dla SQLite bez znaczenia.
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {'pool_pre_ping': True}
app.config['MAX_CONTENT_LENGTH'] = int(
    os.environ.get('MAX_UPLOAD_MB', DEFAULT_MAX_UPLOAD_MB)) * 1024 * 1024
# SameSite=Lax: przegladarka nie dosle ciasteczka przy POST z obcej strony.
# Endpointy JSON byly chronione ubocznie (wymagaja Content-Type: application/json,
# co wymusza preflight), ale /api/import-csv jest multipart, czyli bylo osiagalne
# CSRF-em. Secure wlaczamy zmienna, bo do .31 wchodzi sie tez po HTTP z LAN-u.
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SECURE'] = os.environ.get('SESSION_COOKIE_SECURE') == '1'

db = SQLAlchemy(app)


@app.errorhandler(HTTPException)
def _blad_http(e):
    """Na /api/ zwracaj JSON — front wszedzie czyta `data.error`.

    Domyslnie Flask oddaje strone HTML, wiec `await r.json()` w JS wybuchalo i
    uzytkownik widzial "Blad polaczenia" zamiast tresci bledu.
    """
    if request.path.startswith('/api/'):
        return jsonify({'error': e.description}), e.code
    return e


@app.errorhandler(413)
def _upload_za_duzy(e):
    """Domyslna odpowiedz 413 to HTML — front na /import-csv czyta JSON."""
    limit_mb = app.config['MAX_CONTENT_LENGTH'] // (1024 * 1024)
    komunikat = f'Plik jest za duzy (limit {limit_mb} MB).'
    if request.path.startswith('/api/'):
        return jsonify({'error': komunikat}), 413
    return komunikat, 413


@event.listens_for(Engine, 'connect')
def _set_sqlite_pragmas(dbapi_connection, connection_record):
    """SQLite: WAL + busy timeout, so concurrent leaders don't hit "database is locked".

    WAL lets readers work while one writer commits (rollback-journal mode locks the
    whole file); busy_timeout makes a blocked writer wait 5s instead of failing at once.
    No-op on any other engine — see docs/postgres_schema.sql for the Postgres migration.
    """
    if not isinstance(dbapi_connection, sqlite3.Connection):
        return
    cursor = dbapi_connection.cursor()
    cursor.execute('PRAGMA journal_mode=WAL')
    cursor.execute('PRAGMA busy_timeout=5000')
    cursor.close()


login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# ── Time helpers ───────────────────────────────────────────────────────────────
# Everything is STORED in UTC (naive). For display we serialize with an explicit
# 'Z' so the browser's `new Date()` parses it as UTC and converts to local, and we
# provide a Jinja filter for server-rendered timestamps (single-site: Europe/Warsaw).
LOCAL_TZ = ZoneInfo('Europe/Warsaw')


def iso_z(dt):
    """Serialize a naive-UTC datetime as an explicit-UTC ISO string (…Z).

    Without the 'Z', `new Date(iso)` in the browser parses a date-time string as
    LOCAL, so a stored UTC value would render unconverted (2h off in PL summer).
    Returns None for falsy input. Use ONLY for datetime columns, not date-only.
    """
    return dt.isoformat() + 'Z' if dt else None


def local_today():
    """Dzisiejsza data w Europe/Warsaw — niezaleznie od strefy serwera.

    `date.today()` zwraca date wedlug TZ procesu: w kontenerze bez TZ jest to
    UTC, wiec miedzy 00:00 a 02:00 czasu polskiego oddawalo jeszcze wczoraj.
    """
    return datetime.now(LOCAL_TZ).date()


def utc_to_local(dt):
    """Naive UTC -> aware datetime w Europe/Warsaw."""
    return dt.replace(tzinfo=timezone.utc).astimezone(LOCAL_TZ)


def local_day_bounds(d):
    """Granice doby lokalnej `d` jako naive UTC: [start, koniec).

    Timestampy trzymamy jako naive UTC, a doba pracy jest liczona po Warszawie
    (dwie zmiany, druga pracuje przez polnoc). Polnoc lokalna to 22:00 UTC
    latem i 23:00 UTC zima — DST liczy ZoneInfo, nie staly offset.
    """
    start = datetime.combine(d, time.min, tzinfo=LOCAL_TZ)
    end = datetime.combine(d + timedelta(days=1), time.min, tzinfo=LOCAL_TZ)
    to_naive_utc = lambda dt: dt.astimezone(timezone.utc).replace(tzinfo=None)  # noqa: E731
    return to_naive_utc(start), to_naive_utc(end)


@app.template_filter('localdt')
def localdt_filter(dt, fmt='%d.%m.%Y %H:%M'):
    """Jinja filter: render a naive-UTC datetime in Europe/Warsaw (DST-correct)."""
    if not dt:
        return '—'
    return utc_to_local(dt).strftime(fmt)

# ══════════════════════════════════════════════════════════════════════════════
#  MODELS
# ══════════════════════════════════════════════════════════════════════════════

class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(100), unique=True, nullable=False)
    display_name = db.Column(db.String(150), nullable=False)
    barcode_id = db.Column(db.String(100), unique=True, nullable=True)
    password_hash = db.Column(db.String(200), nullable=True)  # only leaders/admins
    role = db.Column(db.String(20), default='operator')  # operator | leader | admin
    is_active_user = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        if not self.password_hash:
            return False
        return check_password_hash(self.password_hash, password)

    def to_dict(self):
        return {
            'id': self.id,
            'username': self.username,
            'display_name': self.display_name,
            'barcode_id': self.barcode_id,
            'role': self.role,
            'is_active_user': self.is_active_user
        }


class Activity(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)
    sort_order = db.Column(db.Integer, default=0)
    is_active = db.Column(db.Boolean, default=True)

    def to_dict(self):
        return {
            'id': self.id,
            'name': self.name,
            'sort_order': self.sort_order,
            'is_active': self.is_active
        }


class Shift(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, nullable=False)
    shift_number = db.Column(db.Integer, nullable=False)  # 1 or 2
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    __table_args__ = (
        db.UniqueConstraint('date', 'shift_number', name='uq_shift_date_number'),
    )

    def to_dict(self):
        return {
            'id': self.id,
            'date': self.date.isoformat(),
            'shift_number': self.shift_number
        }


class ShiftAttendance(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    shift_id = db.Column(db.Integer, db.ForeignKey('shift.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    scanned_at = db.Column(db.DateTime, default=datetime.utcnow)

    shift = db.relationship('Shift', backref=db.backref('attendances', lazy=True))
    user = db.relationship('User', backref=db.backref('attendances', lazy=True))

    __table_args__ = (
        db.UniqueConstraint('shift_id', 'user_id', name='uq_attendance_shift_user'),
    )

    def to_dict(self):
        return {
            'id': self.id,
            'shift_id': self.shift_id,
            'user_id': self.user_id,
            'user': self.user.to_dict(),
            'scanned_at': iso_z(self.scanned_at)
        }


class ActivityAssignment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    shift_id = db.Column(db.Integer, db.ForeignKey('shift.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    activity_id = db.Column(db.Integer, db.ForeignKey('activity.id'), nullable=False)
    is_suggestion = db.Column(db.Boolean, default=False)
    assigned_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    assigned_at = db.Column(db.DateTime, default=datetime.utcnow)

    shift = db.relationship('Shift', backref=db.backref('assignments', lazy=True))
    user = db.relationship('User', foreign_keys=[user_id],
                           backref=db.backref('assignments', lazy=True))
    activity = db.relationship('Activity', backref=db.backref('assignments', lazy=True))
    assigner = db.relationship('User', foreign_keys=[assigned_by])

    def to_dict(self):
        return {
            'id': self.id,
            'shift_id': self.shift_id,
            'user_id': self.user_id,
            'activity_id': self.activity_id,
            'is_suggestion': self.is_suggestion,
            'user': self.user.to_dict(),
            'activity': self.activity.to_dict()
        }


class DailyStat(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    shift_id = db.Column(db.Integer, db.ForeignKey('shift.id'), nullable=False)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    activity_id = db.Column(db.Integer, db.ForeignKey('activity.id'), nullable=False)
    quantity = db.Column(db.Integer, default=0)
    note = db.Column(db.String(300), nullable=True)

    # Audit fields
    entered_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    entered_at = db.Column(db.DateTime, default=datetime.utcnow)
    modified_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    modified_at = db.Column(db.DateTime, nullable=True)

    shift = db.relationship('Shift', backref=db.backref('stats', lazy=True))
    user = db.relationship('User', foreign_keys=[user_id],
                           backref=db.backref('stats', lazy=True))
    activity = db.relationship('Activity', backref=db.backref('stats', lazy=True))
    entered_by_user = db.relationship('User', foreign_keys=[entered_by])
    modified_by_user = db.relationship('User', foreign_keys=[modified_by])

    __table_args__ = (
        db.UniqueConstraint('shift_id', 'user_id', 'activity_id',
                            name='uq_stat_shift_user_activity'),
    )

    def to_dict(self):
        return {
            'id': self.id,
            'shift_id': self.shift_id,
            'user_id': self.user_id,
            'activity_id': self.activity_id,
            'quantity': self.quantity,
            'note': self.note,
            'entered_by': self.entered_by,
            'entered_at': iso_z(self.entered_at),
            'modified_by': self.modified_by,
            'modified_at': iso_z(self.modified_at),
            'user': self.user.to_dict(),
            'activity': self.activity.to_dict()
        }


class CountryMapping(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    country = db.Column(db.String(150), nullable=False)
    innenauftrag = db.Column(db.String(100), nullable=False)

    def to_dict(self):
        return {
            'id': self.id,
            'country': self.country,
            'innenauftrag': self.innenauftrag
        }


STAT_CATEGORIES = [
    'labelling_on', 'labelling_tvl', 'labelling_try', 'textile',
    'accessoire', 'sunglasses', 'card_facture', 'labelling_polybag',
    'sorting', 'carton_labeling'
]

STAT_CATEGORY_LABELS = {
    'labelling_on': 'Labelling on',
    'labelling_tvl': 'Labelling tvl',
    'labelling_try': 'Labelling try',
    'textile': 'Textile',
    'accessoire': 'accessoire',
    'sunglasses': 'Sunglasses',
    'card_facture': 'Card facture',
    'labelling_polybag': 'Labelling polybag',
    'sorting': 'Sorting',
    'carton_labeling': 'Carton labeling'
}


def empty_category_data():
    """Return default empty category_data dict."""
    return {cat: {'amount': 0, 'cost': 0.0} for cat in STAT_CATEGORIES}


class ImportedCarton(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    barcode = db.Column(db.String(100), unique=True, nullable=False)
    land = db.Column(db.String(150), nullable=True)
    stueckzahl = db.Column(db.Integer, default=0)
    kategorie = db.Column(db.String(100), nullable=True)
    ziel_datum = db.Column(db.Date, nullable=True)
    uebergabe_nr = db.Column(db.String(100), nullable=True)
    country_mapping_id = db.Column(db.Integer, db.ForeignKey('country_mapping.id'), nullable=True)
    imported_at = db.Column(db.DateTime, default=datetime.utcnow)
    imported_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    processed_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    processed_at = db.Column(db.DateTime, nullable=True)
    double_rate = db.Column(db.Boolean, default=False)
    scan_start_at = db.Column(db.DateTime, nullable=True)
    scan_start_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    scan_end_at = db.Column(db.DateTime, nullable=True)
    scan_end_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    added_manually = db.Column(db.Boolean, default=False)
    modified_at = db.Column(db.DateTime, nullable=True)
    modified_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)

    __table_args__ = (
        db.Index('ix_carton_ziel_datum',    'ziel_datum'),
        db.Index('ix_carton_uebergabe_nr',  'uebergabe_nr'),
        db.Index('ix_carton_processed_by',  'processed_by'),
        db.Index('ix_carton_land',          'land'),
        db.Index('ix_carton_imported_at',   'imported_at'),
    )

    country_mapping = db.relationship('CountryMapping', backref=db.backref('cartons', lazy=True))
    processed_by_user = db.relationship('User', foreign_keys=[processed_by])
    scan_start_by_user = db.relationship('User', foreign_keys=[scan_start_by])
    scan_end_by_user = db.relationship('User', foreign_keys=[scan_end_by])
    imported_by_user = db.relationship('User', foreign_keys=[imported_by])
    modified_by_user = db.relationship('User', foreign_keys=[modified_by])

    def processing_seconds(self):
        if self.scan_start_at and self.scan_end_at:
            return int((self.scan_end_at - self.scan_start_at).total_seconds())
        return None

    def to_dict(self):
        return {
            'id': self.id,
            'barcode': self.barcode,
            'land': self.land,
            'stueckzahl': self.stueckzahl,
            'kategorie': self.kategorie,
            'ziel_datum': self.ziel_datum.isoformat() if self.ziel_datum else None,
            'uebergabe_nr': self.uebergabe_nr,
            'country_mapping_id': self.country_mapping_id,
            'processed_by': self.processed_by,
            'double_rate': self.double_rate or False,
            'processed_by_name': self.processed_by_user.display_name if self.processed_by_user else None,
            'processed_at': iso_z(self.processed_at),
            'scan_start_at': iso_z(self.scan_start_at),
            'scan_start_by_name': self.scan_start_by_user.display_name if self.scan_start_by_user else None,
            'scan_end_at': iso_z(self.scan_end_at),
            'scan_end_by_name': self.scan_end_by_user.display_name if self.scan_end_by_user else None,
            'processing_seconds': self.processing_seconds(),
            'added_manually': self.added_manually or False,
            'imported_by_login': self.imported_by_user.username if self.imported_by_user else None,
            'modified_at': iso_z(self.modified_at),
            'modified_by_login': self.modified_by_user.username if self.modified_by_user else None,
        }


class Forecast(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, nullable=False, unique=True)
    quantity = db.Column(db.Integer, default=0)
    notes = db.Column(db.String(500), nullable=True)
    created_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, nullable=True)
    updated_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)

    created_by_user = db.relationship('User', foreign_keys=[created_by])
    updated_by_user = db.relationship('User', foreign_keys=[updated_by])


class GeneralStat(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    loading_date = db.Column(db.Date, nullable=False)
    week_number = db.Column(db.Integer, nullable=False)
    list_id = db.Column(db.String(100), nullable=False)
    country_of_destination = db.Column(db.String(150), nullable=True)
    country_ledger = db.Column(db.String(150), nullable=False)
    amounts = db.Column(db.Integer, default=0)
    category_data = db.Column(db.Text, default='{}')
    double_rate_category_data = db.Column(db.Text, default='{}')  # yellow row: manual per-category amounts for double-rate packages
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, nullable=True)
    updated_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)

    __table_args__ = (
        db.UniqueConstraint('list_id', 'country_ledger', 'loading_date',
                            name='uq_general_stat'),
        db.Index('ix_gstat_loading_date', 'loading_date'),
        db.Index('ix_gstat_list_id',      'list_id'),
    )

    def get_category_data(self):
        try:
            return json.loads(self.category_data) if self.category_data else empty_category_data()
        except (json.JSONDecodeError, TypeError):
            return empty_category_data()

    def set_category_data(self, data):
        self.category_data = json.dumps(data)

    def get_double_rate_category_data(self):
        try:
            return json.loads(self.double_rate_category_data) if self.double_rate_category_data else empty_category_data()
        except (json.JSONDecodeError, TypeError):
            return empty_category_data()

    def set_double_rate_category_data(self, data):
        self.double_rate_category_data = json.dumps(data)

    def to_dict(self):
        """Serializacja linii razem z kosztem policzonym ze stawek jej miesiaca.

        Front nie zapisuje kosztow — sa wyliczane z `CostMapping` przy kazdym
        odczycie. Stawki idą przez `rates_for()` (cache na zadanie), bo inaczej
        kazdy wiersz odpalalby wlasne SELECT-y (N+1 na calej liscie).
        """
        cd = self.get_category_data()
        rates = rates_for(self.loading_date.year, self.loading_date.month)

        total_cost = 0.0
        for cat, data in cd.items():
            amt = data.get('amount', 0)
            rate = rates.get(cat, 0.0)
            data['computed_cost'] = amt * rate
            total_cost += data['computed_cost']

        return {
            'id': self.id,
            'loading_date': self.loading_date.isoformat(),
            'week_number': self.week_number,
            'list_id': self.list_id,
            'country_of_destination': self.country_of_destination,
            'country_ledger': self.country_ledger,
            'amounts': self.amounts,
            'category_data': cd,
            'total_cost': total_cost,
        }


class CostMapping(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    year = db.Column(db.Integer, nullable=False)
    month = db.Column(db.Integer, nullable=False)
    rates_data = db.Column(db.Text, default='{}')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(db.DateTime, nullable=True)
    updated_by = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)

    __table_args__ = (
        db.UniqueConstraint('year', 'month', name='uq_cost_mapping_ym'),
    )

    def get_rates_data(self):
        try:
            return json.loads(self.rates_data) if self.rates_data else {}
        except (json.JSONDecodeError, TypeError):
            return {}

    def set_rates_data(self, data):
        self.rates_data = json.dumps(data)
        
    def to_dict(self):
        return {
            'id': self.id,
            'year': self.year,
            'month': self.month,
            'rates_data': self.get_rates_data()
        }


def rates_for(year, month):
    """Stawki kategorii na dany miesiac, z cache na jedno zadanie.

    `GeneralStat.to_dict()` jest wolane per wiersz, a stawek jest kilka na rok —
    bez cache lista statystyk robila jedno SELECT na wiersz.
    """
    klucz = (year, month)
    cache = g.get('_rates_cache') if has_request_context() else None
    if cache is None:
        cache = {}
        if has_request_context():
            g._rates_cache = cache
    if klucz not in cache:
        mapping = CostMapping.query.filter_by(year=year, month=month).first()
        cache[klucz] = mapping.get_rates_data() if mapping else {}
    return cache[klucz]


class WorkerTimeEvent(db.Model):
    id           = db.Column(db.Integer, primary_key=True)
    user_id      = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    shift_id     = db.Column(db.Integer, db.ForeignKey('shift.id'), nullable=False)
    event_type   = db.Column(db.String(20), nullable=False)  # patrz EVENT_TYPES
    timestamp    = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    recorded_by  = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=True)
    is_manual    = db.Column(db.Boolean, default=False)
    note         = db.Column(db.String(300), nullable=True)

    __table_args__ = (
        db.Index('ix_wte_user_shift', 'user_id', 'shift_id'),
        db.Index('ix_wte_shift_id',   'shift_id'),
    )

    user     = db.relationship('User', foreign_keys=[user_id],
                               backref=db.backref('time_events', lazy=True))
    shift    = db.relationship('Shift', backref=db.backref('time_events', lazy=True))
    recorder = db.relationship('User', foreign_keys=[recorded_by])

    def to_dict(self):
        return {
            'id':             self.id,
            'user_id':        self.user_id,
            'user_name':      self.user.display_name,
            'shift_id':       self.shift_id,
            'event_type':     self.event_type,
            'timestamp':      iso_z(self.timestamp),
            'is_manual':      self.is_manual,
            'note':           self.note or '',
            'recorded_by_name': self.recorder.display_name if self.recorder else None,
        }


EVENT_TYPES = ('break_start', 'break_end', 'work_end')


class AppSetting(db.Model):
    """Generic key/value store for app-wide configurable settings."""
    key = db.Column(db.String(100), primary_key=True)
    value = db.Column(db.String(500), nullable=True)


# Defaults for known settings (used when a key is not yet stored)
SETTING_DEFAULTS = {
    'break_threshold_minutes': '30',
}


def get_setting(key, default=None):
    """Return a stored setting value, else its known default, else `default`."""
    row = AppSetting.query.get(key)
    if row is not None and row.value is not None:
        return row.value
    return SETTING_DEFAULTS.get(key, default)


def get_setting_int(key, default=0):
    try:
        return int(get_setting(key, default))
    except (TypeError, ValueError):
        return default


def set_setting(key, value):
    row = AppSetting.query.get(key)
    if row is None:
        row = AppSetting(key=key, value=str(value))
        db.session.add(row)
    else:
        row.value = str(value)


# ══════════════════════════════════════════════════════════════════════════════
#  AUTH HELPERS
# ══════════════════════════════════════════════════════════════════════════════

ROLES = ('operator', 'leader', 'admin')
# Role z dostepem do logowania i ekranow — konta uprzywilejowane. Zakladac je,
# zmieniac im role/haslo i dezaktywowac moze wylacznie admin.
PRIVILEGED_ROLES = ('leader', 'admin')


@login_manager.user_loader
def load_user(user_id):
    """Zwraca None dla konta nieaktywnego, wiec dezaktywacja wyrzuca tez z
    trwajacej sesji (nie tylko blokuje kolejne logowanie)."""
    try:
        user = db.session.get(User, int(user_id))
    except (TypeError, ValueError):
        return None
    if user is None or not user.is_active_user:
        return None
    return user


def acting_as_admin():
    return current_user.is_authenticated and current_user.role == 'admin'


def active_admin_count(exclude_id=None):
    """Ilu aktywnych adminow zostanie, jesli pominac `exclude_id`."""
    q = User.query.filter_by(role='admin', is_active_user=True)
    if exclude_id is not None:
        q = q.filter(User.id != exclude_id)
    return q.count()


def leader_required(f):
    """Decorator: requires leader or admin role."""
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if current_user.role not in ('leader', 'admin'):
            flash('Brak uprawnień.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    """Decorator: requires admin role."""
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if current_user.role != 'admin':
            flash('Brak uprawnień.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated


# ══════════════════════════════════════════════════════════════════════════════
#  HELPER: get or create shift
# ══════════════════════════════════════════════════════════════════════════════

def json_body():
    """Cialo JSON jako dict. Brak ciala albo zly Content-Type -> {}, zeby
    walidacja pol dala czytelne 400 zamiast 415/500."""
    return request.get_json(silent=True) or {}


def parse_shift_number(value, default=1):
    """Numer zmiany: 1 albo 2. Cokolwiek innego -> `default`."""
    try:
        numer = int(value)
    except (TypeError, ValueError):
        return default
    return numer if numer in (1, 2) else default


def parse_date(value, pole='date'):
    """Data ISO z requestu. Puste -> dzis lokalnie, zle sformatowana -> 400.

    Dotad `date.fromisoformat()` lecialo bez zabezpieczenia i literowka w URL-u
    konczyla sie 500-ka w logach.
    """
    if not value:
        return local_today()
    try:
        return date.fromisoformat(str(value).strip())
    except ValueError:
        abort(400, f'Nieprawidłowa data w "{pole}" (oczekiwano YYYY-MM-DD).')


def require_int(value, pole):
    """Liczba calkowita albo 400 z nazwa pola."""
    try:
        return int(value)
    except (TypeError, ValueError):
        abort(400, f'Pole "{pole}" musi być liczbą całkowitą.')


def get_or_create_shift(shift_date, shift_number):
    """Get existing shift or create new one."""
    shift = Shift.query.filter_by(date=shift_date, shift_number=shift_number).first()
    if not shift:
        shift = Shift(date=shift_date, shift_number=shift_number)
        db.session.add(shift)
        db.session.commit()
    return shift


# ══════════════════════════════════════════════════════════════════════════════
#  PAGE ROUTES
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/')
def index():
    if current_user.is_authenticated:
        return redirect(url_for('scanner', shift_number=1))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        user = User.query.filter_by(username=username).first()
        if (user and user.is_active_user and user.role in PRIVILEGED_ROLES
                and user.check_password(password)):
            login_user(user)
            next_page = request.args.get('next')
            return redirect(next_page or url_for('index'))
        flash('Nieprawidłowy login lub hasło.', 'error')
    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))


@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    if request.method == 'POST':
        current = request.form.get('current_password', '')
        new = request.form.get('new_password', '')
        confirm = request.form.get('confirm_password', '')
        if not current_user.check_password(current):
            flash('Aktualne hasło jest nieprawidłowe.', 'error')
        elif len(new) < 6:
            flash('Nowe hasło musi mieć co najmniej 6 znaków.', 'error')
        elif new != confirm:
            flash('Nowe hasła nie są identyczne.', 'error')
        elif new == current:
            flash('Nowe hasło musi różnić się od aktualnego.', 'error')
        else:
            current_user.set_password(new)
            db.session.commit()
            flash('Hasło zostało zmienione.', 'success')
            return redirect(url_for('profile'))
    return render_template('profile.html')


@app.route('/scanner/<int:shift_number>')
@leader_required
def scanner(shift_number):
    if shift_number not in (1, 2):
        shift_number = 1
    today = local_today()
    shift = get_or_create_shift(today, shift_number)
    attendances = ShiftAttendance.query.filter_by(shift_id=shift.id)\
        .order_by(ShiftAttendance.scanned_at.desc()).all()
    return render_template('scanner.html',
                           shift_number=shift_number,
                           shift=shift,
                           attendances=attendances,
                           today=today)


@app.route('/assignment')
@leader_required
def assignment():
    activities = Activity.query.filter_by(is_active=True)\
        .order_by(Activity.sort_order).all()
    return render_template('assignment.html', activities=activities)


@app.route('/data-entry')
@leader_required
def data_entry():
    activities = Activity.query.filter_by(is_active=True)\
        .order_by(Activity.sort_order).all()
    return render_template('data_entry.html', activities=activities)


@app.route('/stats')
@leader_required
def stats():
    users = User.query.filter_by(is_active_user=True)\
        .order_by(User.display_name).all()
    activities = Activity.query.filter_by(is_active=True)\
        .order_by(Activity.sort_order).all()
    return render_template('stats.html', users=users, activities=activities)


@app.route('/admin/activities')
@admin_required
def admin_activities():
    activities = Activity.query.order_by(Activity.sort_order).all()
    return render_template('admin_activities.html', activities=activities)


@app.route('/admin/users')
@leader_required
def admin_users():
    users = User.query.order_by(User.display_name).all()
    return render_template('admin_users.html', users=users)


@app.route('/admin/panel')
@admin_required
def admin_panel():
    return render_template('admin_panel.html')


@app.route('/admin/country-mapping')
@admin_required
def admin_country_mapping():
    mappings = CountryMapping.query.order_by(CountryMapping.country).all()
    return render_template('admin_country_mapping.html', mappings=mappings)


@app.route('/admin/cost-mapping')
@admin_required
def admin_cost_mapping():
    return render_template('admin_cost_mapping.html', categories=STAT_CATEGORIES, category_labels=STAT_CATEGORY_LABELS)


@app.route('/admin/settings')
@admin_required
def admin_settings():
    return render_template('admin_settings.html',
                           break_threshold=get_setting_int('break_threshold_minutes', 30))


@app.route('/api/settings', methods=['PUT'])
@admin_required
def api_settings_update():
    data = json_body()
    if 'break_threshold_minutes' in data:
        try:
            val = int(data['break_threshold_minutes'])
        except (TypeError, ValueError):
            return jsonify({'error': 'Próg przerwy musi być liczbą całkowitą.'}), 400
        if val < 1:
            return jsonify({'error': 'Próg przerwy musi być większy od zera.'}), 400
        set_setting('break_threshold_minutes', val)
    db.session.commit()
    return jsonify({
        'message': 'Zapisano ustawienia.',
        'break_threshold_minutes': get_setting_int('break_threshold_minutes', 30),
    }), 200


@app.route('/import-csv')
@admin_required
def import_csv_page():
    return render_template('import_csv.html')

@app.route('/paczki')
@leader_required
def paczki_view():
    date_from_str = request.args.get('date_from', '')
    date_to_str = request.args.get('date_to', '')
    barcode = request.args.get('barcode', '').strip()
    land = request.args.get('land', '').strip()
    page = request.args.get('page', 1, type=int)
    
    query = ImportedCarton.query
    
    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            query = query.filter(ImportedCarton.ziel_datum >= date_from)
        except ValueError:
            pass
            
    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            query = query.filter(ImportedCarton.ziel_datum <= date_to)
        except ValueError:
            pass
            
    if barcode:
        query = query.filter(ImportedCarton.barcode.ilike(f'%{barcode}%'))
        
    if land:
        query = query.filter(ImportedCarton.land.ilike(f'%{land}%'))

    pagination = query.order_by(ImportedCarton.imported_at.desc()).paginate(page=page, per_page=100, error_out=False)
    users = User.query.filter_by(is_active_user=True).order_by(User.display_name).all()
    country_mappings = CountryMapping.query.order_by(CountryMapping.country).all()

    return render_template('paczki.html',
                           pagination=pagination,
                           items=pagination.items,
                           date_from=date_from_str,
                           date_to=date_to_str,
                           barcode=barcode,
                           land=land,
                           users=users,
                           country_mappings=country_mappings)


@app.route('/general-stats/export')
@admin_required
def general_stats_export():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import make_response

    date_from_str = request.args.get('date_from', '')
    date_to_str = request.args.get('date_to', '')

    query = GeneralStat.query
    if date_from_str:
        try:
            query = query.filter(GeneralStat.loading_date >= datetime.strptime(date_from_str, '%Y-%m-%d').date())
        except ValueError:
            pass
    if date_to_str:
        try:
            query = query.filter(GeneralStat.loading_date <= datetime.strptime(date_to_str, '%Y-%m-%d').date())
        except ValueError:
            pass
    stats = query.order_by(GeneralStat.loading_date.asc()).all()

    dr_amounts = double_rate_amount_map()

    cost_mappings = CostMapping.query.all()
    rates_by_ym = {(cm.year, cm.month): cm.get_rates_data() for cm in cost_mappings}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Statystyki ogólne'

    # Styles
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill_main = PatternFill('solid', fgColor='3B4A6B')
    header_fill_cat  = PatternFill('solid', fgColor='4F46E5')
    header_fill_dr   = PatternFill('solid', fgColor='C2410C')
    header_fill_tot  = PatternFill('solid', fgColor='166534')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Border(
        left=Side(style='thin', color='555555'),
        right=Side(style='thin', color='555555'),
        top=Side(style='thin', color='555555'),
        bottom=Side(style='thin', color='555555'),
    )
    dr_fill_row = PatternFill('solid', fgColor='FFF3E0')

    # Row 1: main headers
    base_headers = ['Loading date', 'Double Rate', 'Week', 'List-ID',
                    'Country of destination', 'Country ledger', 'Amounts', 'Total Amount']
    cat_cols_start = len(base_headers) + 1

    for ci, h in enumerate(base_headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill_dr if h == 'Double Rate' else header_fill_main
        cell.alignment = center
        cell.border = thin
        ws.merge_cells(start_row=1, start_column=ci, end_row=2, end_column=ci)

    for i, cat in enumerate(STAT_CATEGORIES):
        col = cat_cols_start + i * 2
        label = STAT_CATEGORY_LABELS[cat]
        c = ws.cell(row=1, column=col, value=label)
        c.font = header_font
        c.fill = header_fill_cat
        c.alignment = center
        c.border = thin
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)

    total_col = cat_cols_start + len(STAT_CATEGORIES) * 2
    tc = ws.cell(row=1, column=total_col, value='Total cost (EUR)')
    tc.font = header_font
    tc.fill = header_fill_tot
    tc.alignment = center
    tc.border = thin
    ws.merge_cells(start_row=1, start_column=total_col, end_row=2, end_column=total_col)

    # Row 2: sub-headers for categories
    for i, cat in enumerate(STAT_CATEGORIES):
        col = cat_cols_start + i * 2
        for offset, sub in enumerate(['Amount', 'Cost (EUR)']):
            c = ws.cell(row=2, column=col + offset, value=sub)
            c.font = Font(bold=True, color='FFFFFF', size=9)
            c.fill = header_fill_cat
            c.alignment = center
            c.border = thin

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18

    def write_data_row(ri, s, cd, dr_label, row_fill, amounts_val):
        """Write one data row (normal or yellow double-rate) at Excel row ri."""
        ym = (s.loading_date.year, s.loading_date.month)
        rates = rates_by_ym.get(ym, {})

        total_amount = sum(cd.get(cat, {}).get('amount', 0) for cat in STAT_CATEGORIES)
        total_cost = sum(cd.get(cat, {}).get('amount', 0) * rates.get(cat, 0.0)
                         for cat in STAT_CATEGORIES)

        row_data = [
            s.loading_date.strftime('%d.%m.%Y'),
            dr_label,
            s.week_number,
            s.list_id,
            s.country_of_destination or '',
            s.country_ledger,
            amounts_val,
            total_amount,
        ]

        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = thin
            c.alignment = Alignment(horizontal='center' if ci in (1, 2, 3, 7, 8) else 'left')
            if row_fill:
                c.fill = row_fill
            if ci == 2 and dr_label:
                c.font = Font(bold=True, color='C2410C')

        for i, cat in enumerate(STAT_CATEGORIES):
            col = cat_cols_start + i * 2
            amt = cd.get(cat, {}).get('amount', 0)
            cost = amt * rates.get(cat, 0.0)
            for offset, val in enumerate([amt, round(cost, 2)]):
                c = ws.cell(row=ri, column=col + offset, value=val)
                c.border = thin
                c.alignment = Alignment(horizontal='right')
                if row_fill:
                    c.fill = row_fill

        tc = ws.cell(row=ri, column=total_col, value=round(total_cost, 2))
        tc.border = thin
        tc.alignment = Alignment(horizontal='right')
        tc.font = Font(bold=True)
        if row_fill:
            tc.fill = row_fill

    # Data rows — normal row, then a yellow double-rate row when the line has
    # double-rate packages. The doubling is emergent: the same packages are
    # billed once in the normal row and again (×1) in the yellow row.
    ri = 3
    for s in stats:
        write_data_row(ri, s, s.get_category_data(), '', None, s.amounts)
        ri += 1

        dr_amount = dr_amounts.get((s.list_id, s.country_ledger, s.loading_date), 0)
        if dr_amount > 0:
            write_data_row(ri, s, s.get_double_rate_category_data(),
                           'DOUBLE RATE', dr_fill_row, dr_amount)
            ri += 1

    # Column widths
    col_widths = [12, 10, 6, 14, 22, 18, 9, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    for i in range(len(STAT_CATEGORIES) * 2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(cat_cols_start + i)].width = 9
    ws.column_dimensions[openpyxl.utils.get_column_letter(total_col)].width = 14

    ws.freeze_panes = 'A3'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    fname = f"statystyki_{date_from_str or 'all'}_{date_to_str or 'all'}.xlsx"
    response = make_response(output.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


def double_rate_amount_map():
    """Sum of Stückzahl of double-rate cartons, keyed by
    (uebergabe_nr, land, ziel_datum) == (list_id, country_ledger, loading_date)."""
    from sqlalchemy import func
    rows = db.session.query(
        ImportedCarton.uebergabe_nr,
        ImportedCarton.land,
        ImportedCarton.ziel_datum,
        func.coalesce(func.sum(ImportedCarton.stueckzahl), 0),
    ).filter(ImportedCarton.double_rate.is_(True)).group_by(
        ImportedCarton.uebergabe_nr,
        ImportedCarton.land,
        ImportedCarton.ziel_datum,
    ).all()
    return {(u, l, d): int(s or 0) for u, l, d, s in rows}


@app.route('/general-stats')
@admin_required
def general_stats_page():
    # If parameters not provided, use first and last day of current month
    today = local_today()
    default_from = today.replace(day=1).strftime('%Y-%m-%d')
    # Calculate last day of month
    import calendar
    last_day = calendar.monthrange(today.year, today.month)[1]
    default_to = today.replace(day=last_day).strftime('%Y-%m-%d')

    date_from_str = request.args.get('date_from', default_from)
    date_to_str = request.args.get('date_to', default_to)
    
    query = GeneralStat.query
    
    date_from = None
    date_to = None
    
    if date_from_str:
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
            query = query.filter(GeneralStat.loading_date >= date_from)
        except ValueError:
            pass
            
    if date_to_str:
        try:
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
            query = query.filter(GeneralStat.loading_date <= date_to)
        except ValueError:
            pass

    stats = query.order_by(GeneralStat.loading_date.desc()).all()

    # Ilość double-rate (suma Stückzahl paczek oznaczonych) per linia → żółty wiersz
    dr_amounts = double_rate_amount_map()
    for s in stats:
        s.dr_amount = dr_amounts.get((s.list_id, s.country_ledger, s.loading_date), 0)

    # Przekazanie do szablonu wszystkich mapowań kosztów aby JS miał do nich dostęp przy edycji
    cost_mappings = CostMapping.query.all()
    rates_by_ym = {f"{cm.year}-{cm.month:02d}": cm.get_rates_data() for cm in cost_mappings}
    
    return render_template('general_stats.html',
                           stats=stats,
                           category_labels=STAT_CATEGORY_LABELS,
                           categories=STAT_CATEGORIES,
                           date_from=date_from_str,
                           date_to=date_to_str,
                           rates_by_ym=rates_by_ym)


# ══════════════════════════════════════════════════════════════════════════════
#  API: BARCODE SCANNING
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/scan', methods=['POST'])
@leader_required
def api_scan():
    data = json_body()
    barcode = str(data.get('barcode') or '').strip()
    shift_number = parse_shift_number(data.get('shift_number', 1))
    shift_date = parse_date(data.get('date'))

    if not barcode:
        return jsonify({'error': 'Brak kodu kreskowego.'}), 400

    # Look up user by barcode
    user = User.query.filter_by(barcode_id=barcode, is_active_user=True).first()
    if not user:
        return jsonify({
            'error': 'Nieznany kod kreskowy! Dodaj użytkownika w panelu.'
        }), 404

    # Get or create shift
    shift = get_or_create_shift(shift_date, shift_number)

    # Check if already scanned
    existing = ShiftAttendance.query.filter_by(
        shift_id=shift.id, user_id=user.id
    ).first()
    if existing:
        return jsonify({
            'warning': f'{user.display_name} już zarejestrowany/a na zmianę {shift_number}.',
            'user': user.to_dict(),
            'already_scanned': True
        }), 200

    # Register attendance
    attendance = ShiftAttendance(shift_id=shift.id, user_id=user.id)
    db.session.add(attendance)
    db.session.commit()

    return jsonify({
        'message': f'{user.display_name} zarejestrowany/a na zmianę {shift_number}.',
        'user': user.to_dict(),
        'attendance': attendance.to_dict()
    }), 201


@app.route('/api/scan/<int:attendance_id>', methods=['DELETE'])
@leader_required
def api_unscan(attendance_id):
    attendance = ShiftAttendance.query.get_or_404(attendance_id)
    db.session.delete(attendance)
    db.session.commit()
    return jsonify({'message': 'Usunięto rejestrację.'}), 200


@app.route('/api/shift/attendances', methods=['GET'])
@leader_required
def api_shift_attendances():
    shift_date = request.args.get('date', local_today().isoformat())
    shift_number = parse_shift_number(request.args.get('shift_number', 1))
    shift = Shift.query.filter_by(
        date=parse_date(shift_date),
        shift_number=shift_number
    ).first()
    if not shift:
        return jsonify({'attendances': [], 'shift': None}), 200
    attendances = ShiftAttendance.query.filter_by(shift_id=shift.id)\
        .options(joinedload(ShiftAttendance.user))\
        .order_by(ShiftAttendance.scanned_at.desc()).all()
    return jsonify({
        'attendances': [a.to_dict() for a in attendances],
        'shift': shift.to_dict()
    }), 200


# ══════════════════════════════════════════════════════════════════════════════
#  API: ASSIGNMENT (DRAG & DROP)
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/assignment/data', methods=['GET'])
@leader_required
def api_assignment_data():
    """Get all assignment data for a given date and shift."""
    shift_date = request.args.get('date', local_today().isoformat())
    shift_number = parse_shift_number(request.args.get('shift_number', 1))

    shift = Shift.query.filter_by(
        date=parse_date(shift_date),
        shift_number=shift_number
    ).first()

    activities = Activity.query.filter_by(is_active=True)\
        .order_by(Activity.sort_order).all()

    if not shift:
        return jsonify({
            'shift': None,
            'attendees': [],
            'assignments': [],
            'activities': [a.to_dict() for a in activities]
        }), 200

    # joinedload: to_dict() kazdego wiersza siega po user/activity — bez tego
    # lista robila SELECT na kazdego pracownika i kazda czynnosc osobno.
    attendances = ShiftAttendance.query.filter_by(shift_id=shift.id)\
        .options(joinedload(ShiftAttendance.user)).all()
    assignments = ActivityAssignment.query.filter_by(shift_id=shift.id)\
        .options(joinedload(ActivityAssignment.user),
                 joinedload(ActivityAssignment.activity)).all()

    # Build set of already-assigned user IDs
    assigned_user_ids = {a.user_id for a in assignments}

    return jsonify({
        'shift': shift.to_dict(),
        'attendees': [a.user.to_dict() for a in attendances],
        'assignments': [a.to_dict() for a in assignments],
        'activities': [a.to_dict() for a in activities],
        'unassigned': [a.user.to_dict() for a in attendances
                       if a.user_id not in assigned_user_ids]
    }), 200


@app.route('/api/assignment/suggestions', methods=['GET'])
@leader_required
def api_assignment_suggestions():
    """Generate AI-based assignment suggestions based on user statistics."""
    shift_date = request.args.get('date', local_today().isoformat())
    shift_number = parse_shift_number(request.args.get('shift_number', 1))

    shift = Shift.query.filter_by(
        date=parse_date(shift_date),
        shift_number=shift_number
    ).first()
    if not shift:
        return jsonify({'suggestions': []}), 200

    attendances = ShiftAttendance.query.filter_by(shift_id=shift.id).all()
    attendee_ids = [a.user_id for a in attendances]
    if not attendee_ids:
        return jsonify({'suggestions': []}), 200

    activities = Activity.query.filter_by(is_active=True)\
        .order_by(Activity.sort_order).all()

    # Calculate average quantity per user per activity (last 30 days)
    thirty_days_ago = parse_date(shift_date) - timedelta(days=30)
    stats = db.session.query(
        DailyStat.user_id,
        DailyStat.activity_id,
        func.avg(DailyStat.quantity).label('avg_qty'),
        func.count(DailyStat.id).label('days_worked')
    ).join(Shift).filter(
        DailyStat.user_id.in_(attendee_ids),
        Shift.date >= thirty_days_ago
    ).group_by(DailyStat.user_id, DailyStat.activity_id).all()

    # Build performance map: {(user_id, activity_id): avg_qty}
    perf = {}
    for s in stats:
        perf[(s.user_id, s.activity_id)] = float(s.avg_qty)

    # Greedy assignment: for each activity, pick the best unassigned user
    suggestions = []
    assigned = set()

    # Sort activities by fewest qualified workers first
    activity_scores = []
    for act in activities:
        qualified = sum(1 for uid in attendee_ids
                       if perf.get((uid, act.id), 0) > 0)
        activity_scores.append((qualified, act))
    activity_scores.sort(key=lambda x: x[0])

    for _, act in activity_scores:
        # Rank available users by performance for this activity
        candidates = []
        for uid in attendee_ids:
            if uid not in assigned:
                avg = perf.get((uid, act.id), 0)
                candidates.append((avg, uid))
        candidates.sort(reverse=True)

        if candidates:
            best_avg, best_uid = candidates[0]
            suggestions.append({
                'user_id': best_uid,
                'activity_id': act.id,
                'avg_quantity': best_avg
            })
            assigned.add(best_uid)

    # Assign remaining unassigned users to activities with fewest people
    activity_counts = {act.id: 0 for act in activities}
    for s in suggestions:
        activity_counts[s['activity_id']] = activity_counts.get(s['activity_id'], 0) + 1

    for uid in attendee_ids:
        if uid not in assigned:
            # Find activity with fewest assigned people
            min_act_id = min(activity_counts, key=activity_counts.get)
            suggestions.append({
                'user_id': uid,
                'activity_id': min_act_id,
                'avg_quantity': 0
            })
            activity_counts[min_act_id] += 1
            assigned.add(uid)

    return jsonify({'suggestions': suggestions}), 200


@app.route('/api/assignment/save', methods=['POST'])
@leader_required
def api_assignment_save():
    """Save drag & drop assignments."""
    data = json_body()
    shift_number = parse_shift_number(data.get('shift_number', 1))
    assignments = data.get('assignments') or []
    if not isinstance(assignments, list):
        return jsonify({'error': 'Pole "assignments" musi być listą.'}), 400

    shift = get_or_create_shift(parse_date(data.get('date')), shift_number)

    # Sprawdz FK przed czyszczeniem, zeby bledne cialo nie wyczyscilo przydzialu
    # i nie zostawilo zmiany pustej.
    znani = {u.id for u in User.query.with_entities(User.id).all()}
    czynnosci = {a.id for a in Activity.query.with_entities(Activity.id).all()}
    nowe = []
    for a in assignments:
        if not isinstance(a, dict):
            return jsonify({'error': 'Każde przydzielenie musi być obiektem.'}), 400
        user_id = require_int(a.get('user_id'), 'user_id')
        activity_id = require_int(a.get('activity_id'), 'activity_id')
        if user_id not in znani:
            return jsonify({'error': f'Nieznany pracownik (id {user_id}).'}), 400
        if activity_id not in czynnosci:
            return jsonify({'error': f'Nieznana czynność (id {activity_id}).'}), 400
        nowe.append((user_id, activity_id, bool(a.get('is_suggestion', False))))

    # Clear existing assignments for this shift
    ActivityAssignment.query.filter_by(shift_id=shift.id).delete()

    for user_id, activity_id, is_suggestion in nowe:
        db.session.add(ActivityAssignment(
            shift_id=shift.id,
            user_id=user_id,
            activity_id=activity_id,
            is_suggestion=is_suggestion,
            assigned_by=current_user.id
        ))

    db.session.commit()
    return jsonify({'message': 'Przydzielenie zapisane.'}), 200


# ══════════════════════════════════════════════════════════════════════════════
#  API: DAILY STATS (DATA ENTRY)
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/daily-stats', methods=['GET'])
@leader_required
def api_daily_stats_get():
    """Get stats for a given date and shift."""
    shift_date = request.args.get('date', local_today().isoformat())
    shift_number = parse_shift_number(request.args.get('shift_number', 1))

    shift = Shift.query.filter_by(
        date=parse_date(shift_date),
        shift_number=shift_number
    ).first()
    if not shift:
        return jsonify({'stats': [], 'shift': None}), 200

    # Get assignments for this shift
    assignments = ActivityAssignment.query.filter_by(shift_id=shift.id)\
        .options(joinedload(ActivityAssignment.user),
                 joinedload(ActivityAssignment.activity)).all()
    stats = DailyStat.query.filter_by(shift_id=shift.id)\
        .options(joinedload(DailyStat.user), joinedload(DailyStat.activity)).all()

    return jsonify({
        'shift': shift.to_dict(),
        'assignments': [a.to_dict() for a in assignments],
        'stats': [s.to_dict() for s in stats]
    }), 200


@app.route('/api/daily-stats', methods=['POST'])
@leader_required
def api_daily_stats_save():
    """Save or update daily statistics."""
    data = json_body()
    shift_number = parse_shift_number(data.get('shift_number', 1))
    entries = data.get('entries') or []
    if not isinstance(entries, list):
        return jsonify({'error': 'Pole "entries" musi być listą.'}), 400

    shift = get_or_create_shift(parse_date(data.get('date')), shift_number)

    for entry in entries:
        if not isinstance(entry, dict):
            return jsonify({'error': 'Każdy wpis musi być obiektem.'}), 400
        user_id = require_int(entry.get('user_id'), 'user_id')
        activity_id = require_int(entry.get('activity_id'), 'activity_id')
        quantity = require_int(entry.get('quantity', 0) or 0, 'quantity')

        existing = DailyStat.query.filter_by(
            shift_id=shift.id,
            user_id=user_id,
            activity_id=activity_id
        ).first()

        if existing:
            existing.quantity = quantity
            existing.note = entry.get('note', '')
            existing.modified_by = current_user.id
            existing.modified_at = datetime.utcnow()
        else:
            db.session.add(DailyStat(
                shift_id=shift.id,
                user_id=user_id,
                activity_id=activity_id,
                quantity=quantity,
                note=entry.get('note', ''),
                entered_by=current_user.id
            ))

    db.session.commit()
    return jsonify({'message': 'Statystyki zapisane.'}), 200


@app.route('/api/daily-stats/<int:stat_id>', methods=['PUT'])
@leader_required
def api_daily_stat_update(stat_id):
    """Update a single daily stat (mistake correction)."""
    stat = DailyStat.query.get_or_404(stat_id)
    data = json_body()

    stat.quantity = data.get('quantity', stat.quantity)
    stat.note = data.get('note', stat.note)
    stat.modified_by = current_user.id
    stat.modified_at = datetime.utcnow()

    db.session.commit()
    return jsonify({'message': 'Zaktualizowano.', 'stat': stat.to_dict()}), 200


# ══════════════════════════════════════════════════════════════════════════════
#  API: STATISTICS
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/stats/user/<int:user_id>', methods=['GET'])
@leader_required
def api_stats_user(user_id):
    """Get per-user statistics, grouped by day and month."""
    user = User.query.get_or_404(user_id)
    activity_id = request.args.get('activity_id', type=int)
    date_from = parse_date(request.args.get('date_from'), 'date_from') \
        if request.args.get('date_from') else None
    date_to = parse_date(request.args.get('date_to'), 'date_to') \
        if request.args.get('date_to') else None

    query = db.session.query(
        DailyStat, Shift
    ).join(Shift).filter(DailyStat.user_id == user_id)

    if activity_id:
        query = query.filter(DailyStat.activity_id == activity_id)
    if date_from:
        query = query.filter(Shift.date >= date_from)
    if date_to:
        query = query.filter(Shift.date <= date_to)

    results = query.order_by(Shift.date.desc()).all()

    daily_stats = []
    monthly_agg = {}

    for stat, shift in results:
        daily_stats.append({
            'date': shift.date.isoformat(),
            'shift_number': shift.shift_number,
            'activity': stat.activity.name,
            'activity_id': stat.activity_id,
            'quantity': stat.quantity,
            'note': stat.note,
            'entered_by': stat.entered_by_user.display_name if stat.entered_by_user else None,
            'modified_by': stat.modified_by_user.display_name if stat.modified_by_user else None,
            'modified_at': iso_z(stat.modified_at),
            'stat_id': stat.id
        })

        month_key = shift.date.strftime('%Y-%m')
        act_name = stat.activity.name
        key = (month_key, act_name)
        if key not in monthly_agg:
            monthly_agg[key] = {'total': 0, 'days': 0}
        monthly_agg[key]['total'] += stat.quantity
        monthly_agg[key]['days'] += 1

    # Przetworzone (zakończone) paczki jako syntetyczne pozycje — tylko w widoku "Wszystkie"
    if not activity_id:
        # Grupujemy po DOBIE LOKALNEJ, tak samo jak dashboard. `func.date()`
        # cielo po dacie UTC, wiec paczka zakonczona po lokalnej polnocy trafiala
        # na inny dzien tutaj niz na dashboardzie. Konwersja jest w Pythonie —
        # SQLite i Postgres licza strefy zupelnie inaczej, a zakres jest maly
        # (paczki jednego pracownika w wybranym okresie).
        pq = db.session.query(
            ImportedCarton.scan_end_at,
            ImportedCarton.stueckzahl,
        ).filter(
            ImportedCarton.scan_end_by == user_id,
            ImportedCarton.scan_end_at.isnot(None),
        )
        if date_from:
            pq = pq.filter(ImportedCarton.scan_end_at >= local_day_bounds(date_from)[0])
        if date_to:
            pq = pq.filter(ImportedCarton.scan_end_at < local_day_bounds(date_to)[1])

        per_day = {}
        for scan_end_at, stueckzahl in pq.all():
            d = utc_to_local(scan_end_at).date().isoformat()
            agg = per_day.setdefault(d, {'cnt': 0, 'pcs': 0})
            agg['cnt'] += 1
            agg['pcs'] += stueckzahl or 0

        PKG_METRICS = [('📦 Paczki', lambda cnt, pcs: cnt),
                       ('📦 Paczki (szt.)', lambda cnt, pcs: pcs)]
        for d, agg in per_day.items():
            cnt, pcs = agg['cnt'], agg['pcs']
            for label, metric in PKG_METRICS:
                qty = int(metric(cnt, pcs))
                daily_stats.append({
                    'date': d,
                    'shift_number': None,
                    'activity': label,
                    'activity_id': None,
                    'quantity': qty,
                    'note': '',
                    'entered_by': 'auto (skan paczek)',
                    'modified_by': None,
                    'modified_at': None,
                    'stat_id': None,
                    'is_package': True,
                })
                month_key = d[:7]  # 'YYYY-MM'
                key = (month_key, label)
                if key not in monthly_agg:
                    monthly_agg[key] = {'total': 0, 'days': 0}
                monthly_agg[key]['total'] += qty
                monthly_agg[key]['days'] += 1

    # Wspólne sortowanie malejąco po dacie (paczki wymieszane z czynnościami)
    daily_stats.sort(key=lambda x: x['date'], reverse=True)

    monthly_stats = [
        {
            'month': k[0],
            'activity': k[1],
            'total_quantity': v['total'],
            'days_worked': v['days'],
            'avg_per_day': round(v['total'] / v['days'], 1) if v['days'] > 0 else 0
        }
        for k, v in sorted(monthly_agg.items(), reverse=True)
    ]

    return jsonify({
        'user': user.to_dict(),
        'daily': daily_stats,
        'monthly': monthly_stats
    }), 200


# ══════════════════════════════════════════════════════════════════════════════
#  API: ADMIN — ACTIVITIES
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/activities', methods=['GET'])
@leader_required
def api_activities():
    activities = Activity.query.order_by(Activity.sort_order).all()
    return jsonify([a.to_dict() for a in activities]), 200


@app.route('/api/activities', methods=['POST'])
@admin_required
def api_activity_create():
    data = json_body()
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': 'Nazwa wymagana.'}), 400

    max_order = db.session.query(func.max(Activity.sort_order)).scalar() or 0
    activity = Activity(name=name, sort_order=max_order + 1)
    db.session.add(activity)
    db.session.commit()
    return jsonify(activity.to_dict()), 201


@app.route('/api/activities/<int:activity_id>', methods=['PUT'])
@admin_required
def api_activity_update(activity_id):
    activity = Activity.query.get_or_404(activity_id)
    data = json_body()

    if 'name' in data:
        activity.name = str(data['name'] or '').strip()
    if 'sort_order' in data:
        activity.sort_order = data['sort_order']
    if 'is_active' in data:
        activity.is_active = data['is_active']

    db.session.commit()
    return jsonify(activity.to_dict()), 200


@app.route('/api/activities/<int:activity_id>', methods=['DELETE'])
@admin_required
def api_activity_delete(activity_id):
    activity = Activity.query.get_or_404(activity_id)
    db.session.delete(activity)
    db.session.commit()
    return jsonify({'message': 'Usunięto.'}), 200


@app.route('/api/activities/reorder', methods=['POST'])
@admin_required
def api_activities_reorder():
    data = json_body()
    order = data.get('order', [])
    for i, activity_id in enumerate(order):
        act = Activity.query.get(activity_id)
        if act:
            act.sort_order = i
    db.session.commit()
    return jsonify({'message': 'Kolejność zapisana.'}), 200


# ══════════════════════════════════════════════════════════════════════════════
#  API: ADMIN — USERS
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/users', methods=['GET'])
@leader_required
def api_users():
    users = User.query.order_by(User.display_name).all()
    return jsonify([u.to_dict() for u in users]), 200


@app.route('/api/users', methods=['POST'])
@leader_required
def api_user_create():
    data = json_body()
    username = data.get('username', '').strip()
    display_name = data.get('display_name', '').strip()
    barcode_id = data.get('barcode_id', '').strip()
    role = data.get('role', 'operator')
    password = data.get('password', '')

    if not username or not display_name:
        return jsonify({'error': 'Login i nazwa wyświetlana są wymagane.'}), 400

    if role not in ROLES:
        return jsonify({'error': 'Nieprawidłowa rola.'}), 400

    # Lider zaklada operatorow na zmianie; kont z logowaniem nie moze zalozyc,
    # bo inaczej dorobilby sobie admina.
    if role in PRIVILEGED_ROLES and not acting_as_admin():
        return jsonify({
            'error': 'Tylko administrator może zakładać konta lidera i administratora.'
        }), 403

    if User.query.filter_by(username=username).first():
        return jsonify({'error': 'Taki login już istnieje.'}), 400

    if barcode_id and User.query.filter_by(barcode_id=barcode_id).first():
        return jsonify({'error': 'Taki kod kreskowy już istnieje.'}), 400

    user = User(
        username=username,
        display_name=display_name,
        barcode_id=barcode_id or None,
        role=role
    )
    if role in ('leader', 'admin') and password:
        user.set_password(password)

    db.session.add(user)
    db.session.commit()
    return jsonify(user.to_dict()), 201


@app.route('/api/users/<int:user_id>', methods=['PUT'])
@leader_required
def api_user_update(user_id):
    user = User.query.get_or_404(user_id)
    data = json_body()

    # Lider edytuje wylacznie operatorow i tylko ich dane opisowe. Bez tego
    # moglby podniesc sobie role albo przestawic haslo adminowi i wejsc na jego
    # konto.
    if not acting_as_admin():
        if user.role in PRIVILEGED_ROLES:
            return jsonify({
                'error': 'Konta lidera i administratora może zmieniać tylko administrator.'
            }), 403
        if 'role' in data and data['role'] != user.role:
            return jsonify({'error': 'Tylko administrator może zmieniać rolę.'}), 403
        if data.get('password'):
            return jsonify({'error': 'Tylko administrator może zmieniać hasło.'}), 403

    if 'role' in data and data['role'] not in ROLES:
        return jsonify({'error': 'Nieprawidłowa rola.'}), 400

    # Nie pozwol zdjac uprawnien ostatniemu aktywnemu adminowi — inaczej nikt
    # nie wejdzie juz do panelu.
    degraduje = 'role' in data and user.role == 'admin' and data['role'] != 'admin'
    dezaktywuje = user.role == 'admin' and data.get('is_active_user') is False
    if (degraduje or dezaktywuje) and active_admin_count(exclude_id=user.id) == 0:
        return jsonify({
            'error': 'To ostatnie aktywne konto administratora — najpierw utwórz inne.'
        }), 400

    if 'display_name' in data:
        user.display_name = str(data['display_name'] or '').strip()
    if 'barcode_id' in data:
        new_barcode = str(data['barcode_id'] or '').strip()
        if new_barcode:
            existing = User.query.filter(
                User.barcode_id == new_barcode, User.id != user_id
            ).first()
            if existing:
                return jsonify({'error': 'Taki kod kreskowy już istnieje.'}), 400
            user.barcode_id = new_barcode
        else:
            user.barcode_id = None
    if 'role' in data:
        user.role = data['role']
    if 'is_active_user' in data:
        user.is_active_user = data['is_active_user']
    if 'password' in data and data['password']:
        user.set_password(data['password'])

    db.session.commit()
    return jsonify(user.to_dict()), 200


@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@leader_required
def api_user_delete(user_id):
    user = User.query.get_or_404(user_id)

    if user.role in PRIVILEGED_ROLES and not acting_as_admin():
        return jsonify({
            'error': 'Konta lidera i administratora może dezaktywować tylko administrator.'
        }), 403

    if user.role == 'admin' and active_admin_count(exclude_id=user.id) == 0:
        return jsonify({
            'error': 'To ostatnie aktywne konto administratora — najpierw utwórz inne.'
        }), 400

    user.is_active_user = False  # soft delete — load_user odbiera tez trwajaca sesje
    db.session.commit()
    return jsonify({'message': 'Dezaktywowano.'}), 200


# ══════════════════════════════════════════════════════════════════════════════
#  API: ADMIN — COUNTRY MAPPINGS
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/country-mappings', methods=['GET'])
@admin_required
def api_country_mappings():
    mappings = CountryMapping.query.order_by(CountryMapping.country).all()
    return jsonify([m.to_dict() for m in mappings]), 200


@app.route('/api/country-mappings', methods=['POST'])
@admin_required
def api_country_mapping_create():
    data = json_body()
    country = data.get('country', '').strip()
    innenauftrag = data.get('innenauftrag', '').strip()
    if not country or not innenauftrag:
        return jsonify({'error': 'Kraj i Innenauftrag są wymagane.'}), 400

    mapping = CountryMapping(country=country, innenauftrag=innenauftrag)
    db.session.add(mapping)
    db.session.commit()
    return jsonify(mapping.to_dict()), 201


@app.route('/api/country-mappings/<int:mapping_id>', methods=['PUT'])
@admin_required
def api_country_mapping_update(mapping_id):
    mapping = CountryMapping.query.get_or_404(mapping_id)
    data = json_body()
    if 'country' in data:
        mapping.country = str(data['country'] or '').strip()
    if 'innenauftrag' in data:
        mapping.innenauftrag = str(data['innenauftrag'] or '').strip()
    db.session.commit()
    return jsonify(mapping.to_dict()), 200


@app.route('/api/country-mappings/<int:mapping_id>', methods=['DELETE'])
@admin_required
def api_country_mapping_delete(mapping_id):
    mapping = CountryMapping.query.get_or_404(mapping_id)
    db.session.delete(mapping)
    db.session.commit()
    return jsonify({'message': 'Usunięto mapowanie.'}), 200


# ══════════════════════════════════════════════════════════════════════════════
#  API: CSV IMPORT + GENERAL STATS
# ══════════════════════════════════════════════════════════════════════════════

def detect_csv_encoding(raw_bytes):
    """Try UTF-8 first, then Latin-1."""
    for enc in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
        try:
            raw_bytes.decode(enc)
            return enc
        except (UnicodeDecodeError, LookupError):
            continue
    return 'latin-1'


def normalize_header(h):
    """Normalize CSV header to handle encoding issues."""
    h = h.strip().lower()
    replacements = {
        'st\u00fcckzahl': 'stueckzahl',
        'st\u00fcckel': 'stueckzahl',
        'stã¼ckzahl': 'stueckzahl',
        'stã¼ckel': 'stueckzahl',
        '\u00fcbergabe nr.': 'uebergabe_nr',
        '\u00fcbergabe nr': 'uebergabe_nr',
        'ãœbergabe nr.': 'uebergabe_nr',
        'ãœbergabe nr': 'uebergabe_nr',
        'ziel-datum': 'ziel_datum',
        'land': 'land',
        'barcode': 'barcode',
        'kategorie': 'kategorie',
    }
    for key, val in replacements.items():
        if key in h:
            return val
    return h


def _cell_to_barcode(val):
    """Coerce a cell value to a barcode string.

    Excel stores numeric-looking barcodes as float64 \u2014 a whole number is
    rendered without the trailing '.0' (str(123.0) -> '123'), and leading
    zeros are already lost at the source. CSV always hands us strings.
    """
    if val is None:
        return ''
    if isinstance(val, bool):
        return ''
    if isinstance(val, int):
        return str(val)
    if isinstance(val, float):
        return str(int(val)) if val.is_integer() else str(val)
    return str(val).strip()


def _cell_to_int(val):
    """Coerce a cell value to an int quantity (0 on failure)."""
    if val is None:
        return 0
    if isinstance(val, bool):
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    try:
        return int(float(str(val).strip().replace(',', '')))
    except (ValueError, TypeError):
        return 0


def _cell_to_date(val):
    """Coerce a cell value to a date (None on failure).

    openpyxl returns datetime/date for date cells; CSV returns strings.
    """
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _cell_to_str(val):
    """Coerce a cell value to a trimmed string ('' for None)."""
    if val is None:
        return ''
    return str(val).strip()


# Wiersze importu przetwarzamy partiami: dedup barcode'ow idzie jednym
# zapytaniem `IN (...)` na partie, a nie jednym na wiersz. Rozmiar ograniczony,
# zeby duzy plik nie wchodzil caly do pamieci.
IMPORT_CHUNK = 500


def _partiami(iterable, rozmiar):
    """Dziel iterowalne (moze byc generator) na listy po `rozmiar` elementow."""
    partia = []
    for element in iterable:
        partia.append(element)
        if len(partia) >= rozmiar:
            yield partia
            partia = []
    if partia:
        yield partia


def process_import_rows(rows):
    """Shared import pipeline for CSV and Excel.

    `rows` is an iterable of dicts keyed by NORMALIZED header names
    (barcode, land, stueckzahl, kategorie, ziel_datum, uebergabe_nr).
    Values may be raw strings (CSV) or native types (Excel) \u2014 all field
    coercion is type-aware. Creates ImportedCarton records (deduped by
    barcode) and upserts aggregated GeneralStat rows. Returns the result
    dict for the API response.
    """
    imported = 0
    skipped = 0
    skipped_barcodes = []
    aggregation = {}  # key: (list_id, land, date_str) -> sum of stueckzahl

    # Mapowania krajow to kilkadziesiat rekordow — jeden odczyt zamiast SELECT-a
    # na kazdy wiersz pliku.
    mapowania = {m.innenauftrag: m for m in CountryMapping.query.all()}
    # Barcode'y widziane w TYM imporcie — lapie duplikaty wewnatrz pliku bez
    # polegania na autoflushu.
    widziane = set()

    for partia in _partiami(rows, IMPORT_CHUNK):
        parsed = []
        for row in partia:
            barcode = _cell_to_barcode(row.get('barcode'))
            if not barcode:
                continue
            parsed.append((barcode, row))

        # Jedno zapytanie na partie zamiast jednego na wiersz.
        istniejace = set()
        if parsed:
            kandydaci = [b for b, _ in parsed]
            istniejace = {
                b for (b,) in db.session.query(ImportedCarton.barcode)
                .filter(ImportedCarton.barcode.in_(kandydaci)).all()
            }

        for barcode, row in parsed:
            if barcode in istniejace or barcode in widziane:
                skipped += 1
                skipped_barcodes.append(barcode)
                continue
            widziane.add(barcode)

            land = _cell_to_str(row.get('land'))
            kategorie = _cell_to_str(row.get('kategorie'))
            uebergabe_nr = _cell_to_str(row.get('uebergabe_nr'))
            stueckzahl = _cell_to_int(row.get('stueckzahl'))
            ziel_datum = _cell_to_date(row.get('ziel_datum'))

            mapping = mapowania.get(land)

            carton = ImportedCarton(
                barcode=barcode,
                land=land,
                stueckzahl=stueckzahl,
                kategorie=kategorie,
                ziel_datum=ziel_datum,
                uebergabe_nr=uebergabe_nr,
                country_mapping_id=mapping.id if mapping else None,
                double_rate=bool(row.get('double_rate')),
                added_manually=bool(row.get('added_manually')),
                imported_by=current_user.id
            )
            db.session.add(carton)
            imported += 1

            # Aggregate for GeneralStat
            if uebergabe_nr and ziel_datum:
                agg_key = (uebergabe_nr, land, ziel_datum.isoformat())
                if agg_key not in aggregation:
                    aggregation[agg_key] = {
                        'stueckzahl': 0,
                        'country': mapping.country if mapping else None,
                        'date': ziel_datum
                    }
                aggregation[agg_key]['stueckzahl'] += stueckzahl

    # Create/update GeneralStat entries
    stats_created = 0
    stats_updated = 0
    for (list_id, land_val, date_str), agg in aggregation.items():
        existing = GeneralStat.query.filter_by(
            list_id=list_id,
            country_ledger=land_val,
            loading_date=agg['date']
        ).first()

        if existing:
            existing.amounts += agg['stueckzahl']
            existing.updated_at = datetime.utcnow()
            existing.updated_by = current_user.id
            stats_updated += 1
        else:
            stat = GeneralStat(
                loading_date=agg['date'],
                week_number=agg['date'].isocalendar()[1],
                list_id=list_id,
                country_of_destination=agg['country'],
                country_ledger=land_val,
                amounts=agg['stueckzahl'],
                category_data=json.dumps(empty_category_data())
            )
            db.session.add(stat)
            stats_created += 1

    db.session.commit()

    return {
        'message': f'Zaimportowano {imported} karton\u00f3w.',
        'imported': imported,
        'skipped': skipped,
        'skipped_barcodes': skipped_barcodes[:50],
        'stats_created': stats_created,
        'stats_updated': stats_updated
    }


@app.route('/api/import-csv', methods=['POST'])
@admin_required
def api_import_csv():
    if 'file' not in request.files:
        return jsonify({'error': 'Brak pliku.'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'Brak nazwy pliku.'}), 400

    raw_bytes = file.read()
    encoding = detect_csv_encoding(raw_bytes)
    text = raw_bytes.decode(encoding)

    reader = csv.DictReader(io.StringIO(text), delimiter=';')

    # Normalize headers
    if reader.fieldnames:
        header_map = {h: normalize_header(h) for h in reader.fieldnames}
    else:
        return jsonify({'error': 'Plik CSV nie ma nag\u0142\u00f3wk\u00f3w.'}), 400

    rows = ({header_map.get(k, k): v for k, v in row.items()} for row in reader)
    return jsonify(process_import_rows(rows)), 200


@app.route('/api/import/excel', methods=['POST'])
@admin_required
def api_import_excel():
    import openpyxl

    if 'file' not in request.files:
        return jsonify({'error': 'Brak pliku.'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'Brak nazwy pliku.'}), 400

    raw_bytes = file.read()
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(raw_bytes), data_only=True, read_only=True
        )
    except Exception as e:
        return jsonify({'error': f'Nie mo\u017cna odczyta\u0107 pliku Excel: {e}'}), 400

    ws = wb.active
    row_iter = ws.iter_rows(values_only=True)

    try:
        header_row = next(row_iter)
    except StopIteration:
        return jsonify({'error': 'Plik Excel jest pusty.'}), 400

    header_map = {}
    for idx, cell in enumerate(header_row):
        raw = str(cell) if cell is not None else ''
        header_map[idx] = normalize_header(raw)

    if not any(header_map.values()):
        return jsonify({'error': 'Plik Excel nie ma nag\u0142\u00f3wk\u00f3w.'}), 400

    def rows_gen():
        for values in row_iter:
            yield {header_map.get(idx, idx): val for idx, val in enumerate(values)}

    result = jsonify(process_import_rows(rows_gen())), 200
    wb.close()
    return result


@app.route('/api/general-stats', methods=['GET'])
@admin_required
def api_general_stats():
    """Lista linii rozliczeniowych. Opcjonalne `date_from` / `date_to`
    (YYYY-MM-DD) — bez nich zwraca wszystko, jak dotad."""
    query = GeneralStat.query
    for param, warunek in (('date_from', GeneralStat.loading_date.__ge__),
                           ('date_to', GeneralStat.loading_date.__le__)):
        wartosc = request.args.get(param, '')
        if wartosc:
            try:
                query = query.filter(warunek(date.fromisoformat(wartosc)))
            except ValueError:
                return jsonify({'error': f'Nieprawidłowa data w {param}.'}), 400
    stats = query.order_by(GeneralStat.loading_date.desc()).all()
    return jsonify([s.to_dict() for s in stats]), 200


@app.route('/api/general-stats/<int:stat_id>', methods=['PUT'])
@admin_required
def api_general_stat_update(stat_id):
    stat = GeneralStat.query.get_or_404(stat_id)
    data = json_body()

    if 'category_data' in data:
        stat.set_category_data(data['category_data'])
    if 'double_rate_category_data' in data:
        stat.set_double_rate_category_data(data['double_rate_category_data'])
    stat.updated_at = datetime.utcnow()
    stat.updated_by = current_user.id

    db.session.commit()
    return jsonify(stat.to_dict()), 200


@app.route('/api/cost-mapping/<int:year>/<int:month>', methods=['GET'])
@admin_required
def api_cost_mapping_get(year, month):
    mapping = CostMapping.query.filter_by(year=year, month=month).first()
    if mapping:
        return jsonify(mapping.get_rates_data()), 200
    return jsonify({}), 200


@app.route('/dashboard')
@leader_required
def dashboard():
    return render_template('dashboard.html')


@app.route('/api/dashboard')
@leader_required
def api_dashboard():
    today = local_today()
    today_start, today_end = local_day_bounds(today)

    # Paczki — ogólne (przetworzona = ZAKOŃCZONA, scan_end)
    total_cartons = ImportedCarton.query.count()
    done_cartons = ImportedCarton.query.filter(ImportedCarton.scan_end_at.isnot(None)).count()
    remaining_cartons = total_cartons - done_cartons

    # Paczki zakończone dziś
    dzis = and_(ImportedCarton.scan_end_at >= today_start,
                ImportedCarton.scan_end_at < today_end)
    done_today = ImportedCarton.query.filter(dzis).count()
    pieces_today = db.session.query(func.sum(ImportedCarton.stueckzahl))\
        .filter(dzis).scalar() or 0

    # Breakdown per pracownik — dziś (kto zakończył paczki)
    worker_rows = db.session.query(
        User.id.label('user_id'), User.display_name,
        func.count(ImportedCarton.id).label('packages'),
        func.sum(ImportedCarton.stueckzahl).label('pieces')
    ).join(ImportedCarton, ImportedCarton.scan_end_by == User.id)\
     .filter(dzis)\
     .group_by(User.id, User.display_name)\
     .order_by(func.count(ImportedCarton.id).desc()).all()

    workers_today = [
        {'name': r.display_name, 'packages': r.packages, 'pieces': int(r.pieces or 0)}
        for r in worker_rows
    ]

    # Obecni na zmianach dziś
    shift1 = Shift.query.filter_by(date=today, shift_number=1).first()
    shift2 = Shift.query.filter_by(date=today, shift_number=2).first()
    present_ids = set()
    if shift1:
        present_ids |= {a.user_id for a in ShiftAttendance.query.filter_by(shift_id=shift1.id).all()}
    if shift2:
        present_ids |= {a.user_id for a in ShiftAttendance.query.filter_by(shift_id=shift2.id).all()}

    # DailyStat dziś — per czynność (sumy)
    daily_rows = db.session.query(
        DailyStat.activity_id,
        func.sum(DailyStat.quantity).label('total_qty'),
    ).join(Shift).filter(Shift.date == today).group_by(DailyStat.activity_id).all()

    act_map = {a.id: a.name for a in Activity.query.all()}
    activities_today = [
        {'activity': act_map.get(r.activity_id, '?'), 'total_qty': int(r.total_qty or 0)}
        for r in sorted(daily_rows, key=lambda r: r.total_qty or 0, reverse=True)
    ]

    # Per pracownik na zmianie — paczki + DailyStat
    pkg_by_user = {r.user_id: {'packages': r.packages, 'pieces': int(r.pieces or 0)}
                   for r in worker_rows}

    daily_per_user = db.session.query(
        DailyStat.user_id,
        DailyStat.activity_id,
        func.sum(DailyStat.quantity).label('qty'),
    ).join(Shift).filter(Shift.date == today).group_by(DailyStat.user_id, DailyStat.activity_id).all()

    stat_by_user = {}
    for r in daily_per_user:
        stat_by_user.setdefault(r.user_id, []).append(
            {'activity': act_map.get(r.activity_id, '?'), 'qty': int(r.qty or 0)}
        )

    # Widoczni: obecni na zmianie LUB ci, którzy dziś zakończyli paczki (nic nie chowamy)
    relevant_ids = present_ids | set(pkg_by_user.keys())
    present_users = User.query.filter(User.id.in_(relevant_ids)).order_by(User.display_name).all() \
        if relevant_ids else []

    per_worker = []
    for u in present_users:
        pkg = pkg_by_user.get(u.id, {'packages': 0, 'pieces': 0})
        per_worker.append({
            'id': u.id,
            'name': u.display_name,
            'packages_today': pkg['packages'],
            'pieces_today': pkg['pieces'],
            'activities': sorted(stat_by_user.get(u.id, []), key=lambda x: x['qty'], reverse=True),
        })

    progress_pct = round(done_cartons / total_cartons * 100, 1) if total_cartons else 0

    return jsonify({
        'total_cartons': total_cartons,
        'done_cartons': done_cartons,
        'remaining_cartons': remaining_cartons,
        'done_today': done_today,
        'pieces_today': int(pieces_today),
        'progress_pct': progress_pct,
        'present_count': len(present_ids),
        'workers_today': workers_today,
        'activities_today': activities_today,
        'per_worker': per_worker,
        'as_of': datetime.now(LOCAL_TZ).strftime('%H:%M:%S'),
    })


@app.route('/api/dashboard/shifts')
@leader_required
def api_dashboard_shifts():
    """Per-shift daily breakdown for a chosen date.

    DailyStat is already shift-tagged (shift_id) → aggregated directly.
    Packages (ImportedCarton) have no shift, so they are attributed by
    ATTENDANCE: a package finished by a worker counts toward the single shift
    that worker was scanned into that day. Workers with no attendance (or in
    both shifts → ambiguous) land in the 'unattributed' bucket so nothing is
    silently dropped and totals stay exact (each package counted once).
    """
    date_str = request.args.get('date', '')
    try:
        target = datetime.strptime(date_str, '%Y-%m-%d').date() if date_str else local_today()
    except ValueError:
        target = local_today()

    day_start, day_end = local_day_bounds(target)

    act_map = {a.id: a.name for a in Activity.query.all()}
    user_map = {u.id: u.display_name for u in User.query.all()}

    # Shifts that day + attendance (user_id -> set of shift_numbers attended)
    shifts = {s.shift_number: s for s in Shift.query.filter_by(date=target).all()}
    user_shifts = {}
    for snum, sh in shifts.items():
        for a in ShiftAttendance.query.filter_by(shift_id=sh.id).all():
            user_shifts.setdefault(a.user_id, set()).add(snum)

    # Packages finished that day, grouped by the worker who ended them
    pkg_rows = db.session.query(
        ImportedCarton.scan_end_by.label('user_id'),
        func.count(ImportedCarton.id).label('packages'),
        func.coalesce(func.sum(ImportedCarton.stueckzahl), 0).label('pieces'),
    ).filter(
        ImportedCarton.scan_end_at >= day_start,
        ImportedCarton.scan_end_at < day_end,
        ImportedCarton.scan_end_by.isnot(None),
    ).group_by(ImportedCarton.scan_end_by).all()

    per_shift_pkg = {1: {'packages': 0, 'pieces': 0}, 2: {'packages': 0, 'pieces': 0}}
    unattributed = {'packages': 0, 'pieces': 0, 'workers': []}

    for r in pkg_rows:
        attended = user_shifts.get(r.user_id, set())
        pkgs, pcs = int(r.packages), int(r.pieces or 0)
        if len(attended) == 1:
            snum = next(iter(attended))
            per_shift_pkg[snum]['packages'] += pkgs
            per_shift_pkg[snum]['pieces'] += pcs
        else:
            unattributed['packages'] += pkgs
            unattributed['pieces'] += pcs
            unattributed['workers'].append({
                'name': user_map.get(r.user_id, '?'),
                'packages': pkgs, 'pieces': pcs,
                'reason': 'brak obecności' if not attended else 'obie zmiany',
            })

    # DailyStat per shift + activity (direct via shift_id)
    ds_rows = db.session.query(
        Shift.shift_number,
        DailyStat.activity_id,
        func.sum(DailyStat.quantity).label('qty'),
    ).join(Shift, DailyStat.shift_id == Shift.id)\
     .filter(Shift.date == target)\
     .group_by(Shift.shift_number, DailyStat.activity_id).all()

    per_shift_acts = {1: [], 2: []}
    for r in ds_rows:
        per_shift_acts.setdefault(r.shift_number, []).append(
            {'activity': act_map.get(r.activity_id, '?'), 'qty': int(r.qty or 0)}
        )

    result_shifts = []
    for snum in (1, 2):
        sh = shifts.get(snum)
        present = ShiftAttendance.query.filter_by(shift_id=sh.id).count() if sh else 0
        result_shifts.append({
            'shift_number': snum,
            'exists': sh is not None,
            'present_count': present,
            'packages': per_shift_pkg[snum]['packages'],
            'pieces': per_shift_pkg[snum]['pieces'],
            'activities': sorted(per_shift_acts.get(snum, []), key=lambda x: x['qty'], reverse=True),
        })

    return jsonify({
        'date': target.isoformat(),
        'shifts': result_shifts,
        'unattributed': unattributed,
    })


@app.route('/scan-package')
@leader_required
def scan_package():
    return render_template('scan_package.html')


@app.route('/api/package-lookup', methods=['GET'])
@leader_required
def api_package_lookup():
    """Read-only lookup: sprawdza status paczki (przeskanowana / kto / zakończona) + dane."""
    barcode = (request.args.get('barcode') or '').strip()
    if not barcode:
        return jsonify({'error': 'Brak kodu paczki.'}), 400

    carton = ImportedCarton.query.filter_by(barcode=barcode).first()
    if not carton:
        return jsonify({'error': 'Nieznany kod paczki. Upewnij się, że paczka została zaimportowana.'}), 404

    d = carton.to_dict()
    # Derived status: przeskanowana = przypisana LUB rozpoczęto pomiar czasu
    d['scanned'] = bool(carton.processed_by) or bool(carton.scan_start_at)
    d['scanned_by'] = d.get('processed_by_name') or d.get('scan_start_by_name')
    d['finished'] = bool(carton.scan_end_at)
    return jsonify({'carton': d}), 200


@app.route('/api/packages', methods=['POST'])
@leader_required
def api_package_create():
    """Manually add a single package (for packages missing from CSV imports).

    Runs through the SAME pipeline as CSV/Excel import (process_import_rows),
    so a manual package dedups, aggregates into GeneralStat and bills exactly
    like an imported one. Duplicate barcode → 409; missing required fields or
    bad values → 400.
    """
    from sqlalchemy.exc import IntegrityError

    data = json_body()
    barcode = (data.get('barcode') or '').strip()
    land = (data.get('land') or '').strip()
    kategorie = (data.get('kategorie') or '').strip()
    uebergabe_nr = (data.get('uebergabe_nr') or '').strip()
    ziel_datum_str = (data.get('ziel_datum') or '').strip()
    double_rate = bool(data.get('double_rate'))

    # Required fields (5) — must not be skipped, else the package can't bill
    missing = []
    if not barcode:
        missing.append('Barcode')
    if not land:
        missing.append('Land')
    if not uebergabe_nr:
        missing.append('Übergabe Nr.')
    if not ziel_datum_str:
        missing.append('Ziel-Datum')
    if missing:
        return jsonify({'error': 'Brakuje wymaganych pól: ' + ', '.join(missing)}), 400

    # Quantity — must be a positive integer
    try:
        stueckzahl = int(data.get('stueckzahl'))
    except (TypeError, ValueError):
        return jsonify({'error': 'Stückzahl musi być liczbą całkowitą.'}), 400
    if stueckzahl <= 0:
        return jsonify({'error': 'Stückzahl musi być większe od zera.'}), 400

    # Date — must parse
    ziel_datum = _cell_to_date(ziel_datum_str)
    if ziel_datum is None:
        return jsonify({'error': 'Nieprawidłowy format daty Ziel-Datum.'}), 400

    # Duplicate barcode → 409 (pre-check; race handled by IntegrityError below)
    if ImportedCarton.query.filter_by(barcode=barcode).first():
        return jsonify({'error': f'Paczka o barcode "{barcode}" już istnieje.'}), 409

    row = {
        'barcode': barcode,
        'land': land,
        'stueckzahl': stueckzahl,
        'kategorie': kategorie,
        'ziel_datum': ziel_datum,
        'uebergabe_nr': uebergabe_nr,
        'double_rate': double_rate,
        'added_manually': True,
    }

    try:
        result = process_import_rows([row])
    except IntegrityError:
        # Concurrent add of the same barcode slipped past the pre-check
        db.session.rollback()
        return jsonify({'error': f'Paczka o barcode "{barcode}" już istnieje.'}), 409

    # Guard: a silent skip (imported=0) must not read as success
    if result.get('imported') != 1:
        return jsonify({'error': 'Nie udało się dodać paczki.'}), 409

    return jsonify({
        'message': f'Dodano paczkę {barcode}.',
        'stats_created': result.get('stats_created', 0),
        'stats_updated': result.get('stats_updated', 0),
    }), 201


def recompute_general_stat(uebergabe_nr, land, ziel_datum, actor_id=None):
    """Recompute a GeneralStat line's `amounts` from the SUM of its cartons.

    `amounts` is a pure carton aggregate (only written by process_import_rows),
    so recomputing from SUM(stueckzahl) over the (uebergabe_nr, land, ziel_datum)
    group is exact — it avoids delta drift and stays correct when manual and
    imported cartons share a line. Groups without uebergabe_nr or ziel_datum
    never aggregate, so nothing to recompute. If the group emptied to 0, the
    GeneralStat row is kept at amounts=0 (preserves any entered category_data).
    """
    if not (uebergabe_nr and ziel_datum):
        return
    total = db.session.query(
        func.coalesce(func.sum(ImportedCarton.stueckzahl), 0)
    ).filter(
        ImportedCarton.uebergabe_nr == uebergabe_nr,
        ImportedCarton.land == land,
        ImportedCarton.ziel_datum == ziel_datum,
    ).scalar() or 0

    existing = GeneralStat.query.filter_by(
        list_id=uebergabe_nr,
        country_ledger=land,
        loading_date=ziel_datum,
    ).first()

    if existing:
        existing.amounts = total
        existing.updated_at = datetime.utcnow()
        if actor_id:
            existing.updated_by = actor_id
    elif total > 0:
        mapping = CountryMapping.query.filter_by(innenauftrag=land).first()
        stat = GeneralStat(
            loading_date=ziel_datum,
            week_number=ziel_datum.isocalendar()[1],
            list_id=uebergabe_nr,
            country_of_destination=mapping.country if mapping else None,
            country_ledger=land,
            amounts=total,
            category_data=json.dumps(empty_category_data()),
        )
        db.session.add(stat)


@app.route('/api/packages/<int:carton_id>', methods=['PUT'])
@leader_required
def api_package_update(carton_id):
    """Edit a manually-added package; recompute affected GeneralStat lines.

    Only `added_manually` packages are editable (imported ones are read-only →
    403). Changing a group field (uebergabe/land/ziel) moves the carton between
    GeneralStat groups — both the old and new lines are recomputed from carton
    sums so billing stays exact.
    """
    from sqlalchemy.exc import IntegrityError

    carton = ImportedCarton.query.get_or_404(carton_id)
    if not carton.added_manually:
        return jsonify({'error': 'Edytować można tylko paczki dodane ręcznie.'}), 403

    data = json_body()
    barcode = (data.get('barcode') or '').strip()
    land = (data.get('land') or '').strip()
    kategorie = (data.get('kategorie') or '').strip()
    uebergabe_nr = (data.get('uebergabe_nr') or '').strip()
    ziel_datum_str = (data.get('ziel_datum') or '').strip()
    double_rate = bool(data.get('double_rate'))

    missing = []
    if not barcode:
        missing.append('Barcode')
    if not land:
        missing.append('Land')
    if not uebergabe_nr:
        missing.append('Übergabe Nr.')
    if not ziel_datum_str:
        missing.append('Ziel-Datum')
    if missing:
        return jsonify({'error': 'Brakuje wymaganych pól: ' + ', '.join(missing)}), 400

    try:
        stueckzahl = int(data.get('stueckzahl'))
    except (TypeError, ValueError):
        return jsonify({'error': 'Stückzahl musi być liczbą całkowitą.'}), 400
    if stueckzahl <= 0:
        return jsonify({'error': 'Stückzahl musi być większe od zera.'}), 400

    ziel_datum = _cell_to_date(ziel_datum_str)
    if ziel_datum is None:
        return jsonify({'error': 'Nieprawidłowy format daty Ziel-Datum.'}), 400

    if barcode != carton.barcode and ImportedCarton.query.filter_by(barcode=barcode).first():
        return jsonify({'error': f'Paczka o barcode "{barcode}" już istnieje.'}), 409

    # Capture the OLD group before mutating, so we can recompute it after the move
    old_group = (carton.uebergabe_nr, carton.land, carton.ziel_datum)

    mapping = CountryMapping.query.filter_by(innenauftrag=land).first()
    carton.barcode = barcode
    carton.land = land
    carton.stueckzahl = stueckzahl
    carton.kategorie = kategorie
    carton.ziel_datum = ziel_datum
    carton.uebergabe_nr = uebergabe_nr
    carton.double_rate = double_rate
    carton.country_mapping_id = mapping.id if mapping else None
    carton.modified_at = datetime.utcnow()
    carton.modified_by = current_user.id

    new_group = (uebergabe_nr, land, ziel_datum)

    try:
        db.session.flush()  # push the carton UPDATE so recompute SUMs see new values / catch collision
        recompute_general_stat(*old_group, actor_id=current_user.id)
        if new_group != old_group:
            recompute_general_stat(*new_group, actor_id=current_user.id)
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return jsonify({'error': f'Paczka o barcode "{barcode}" już istnieje.'}), 409

    return jsonify({'message': f'Zapisano paczkę {barcode}.', 'carton': carton.to_dict()}), 200


@app.route('/api/packages/<int:carton_id>/assign', methods=['PUT'])
@leader_required
def api_package_reassign(carton_id):
    carton = ImportedCarton.query.get_or_404(carton_id)
    data = json_body()
    user_id = data.get('user_id')
    if user_id is None:
        carton.processed_by = None
        carton.processed_at = None
    else:
        user = User.query.get_or_404(user_id)
        carton.processed_by = user.id
        carton.processed_at = datetime.utcnow()
    db.session.commit()
    return jsonify({'carton': carton.to_dict()}), 200


@app.route('/api/packages/<int:carton_id>/double-rate', methods=['PUT'])
@leader_required
def api_package_double_rate(carton_id):
    carton = ImportedCarton.query.get_or_404(carton_id)
    data = json_body()
    carton.double_rate = bool(data.get('double_rate'))
    db.session.commit()
    return jsonify({'carton': carton.to_dict()}), 200


@app.route('/api/cost-mapping/<int:year>/<int:month>', methods=['PUT', 'POST'])
@admin_required
def api_cost_mapping_save(year, month):
    data = json_body()
    rates = data.get('rates', {})

    mapping = CostMapping.query.filter_by(year=year, month=month).first()
    if mapping:
        mapping.set_rates_data(rates)
        mapping.updated_at = datetime.utcnow()
        mapping.updated_by = current_user.id
    else:
        mapping = CostMapping(
            year=year,
            month=month,
            updated_by=current_user.id
        )
        mapping.set_rates_data(rates)
        db.session.add(mapping)

    db.session.commit()
    return jsonify({'success': True, 'mapping': mapping.to_dict()}), 200


# ══════════════════════════════════════════════════════════════════════════════
#  PACKAGE SCAN TIMES & DOUBLE RATE
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/scan-paczki')
@leader_required
def scan_paczki():
    return render_template('scan_paczki.html')


@app.route('/api/package-time/start', methods=['POST'])
@leader_required
def api_package_time_start():
    data = json_body()
    employee_barcode = (data.get('employee_barcode') or '').strip()
    package_barcode = (data.get('package_barcode') or '').strip()

    if not employee_barcode:
        return jsonify({'error': 'Brak kodu pracownika.'}), 400
    if not package_barcode:
        return jsonify({'error': 'Brak kodu paczki.'}), 400

    user = User.query.filter_by(barcode_id=employee_barcode, is_active_user=True).first()
    if not user:
        return jsonify({'error': 'Nieznany kod pracownika.'}), 404

    carton = ImportedCarton.query.filter_by(barcode=package_barcode).first()
    if not carton:
        return jsonify({'error': 'Nieznany kod paczki.'}), 404

    # Paczka już zakończona — nie można jej rozpocząć ponownie
    if carton.scan_end_at:
        return jsonify({'error': 'Paczka jest już zakończona — nie można jej rozpocząć ponownie.'}), 409

    # Paczka w trakcie (start bez końca) — tylko pracownik, który zaczął, nią zarządza
    if carton.scan_start_at:
        if carton.scan_start_by == user.id:
            return jsonify({
                'message': f'Paczka {package_barcode} już rozpoczęta przez Ciebie — start bez zmian.',
                'carton': carton.to_dict()
            }), 200
        starter = User.query.get(carton.scan_start_by) if carton.scan_start_by else None
        return jsonify({
            'error': f'Paczka już rozpoczęta przez {starter.display_name if starter else "innego pracownika"}. Nie można jej podebrać.'
        }), 409

    # Świeży start (paczka nierozpoczęta)
    carton.scan_start_at = datetime.utcnow()
    carton.scan_start_by = user.id
    db.session.commit()

    return jsonify({
        'message': f'Start zarejestrowany dla paczki {package_barcode} — {user.display_name}.',
        'carton': carton.to_dict()
    }), 200


@app.route('/api/package-time/end', methods=['POST'])
@leader_required
def api_package_time_end():
    data = json_body()
    employee_barcode = (data.get('employee_barcode') or '').strip()
    package_barcode = (data.get('package_barcode') or '').strip()

    if not employee_barcode:
        return jsonify({'error': 'Brak kodu pracownika.'}), 400
    if not package_barcode:
        return jsonify({'error': 'Brak kodu paczki.'}), 400

    user = User.query.filter_by(barcode_id=employee_barcode, is_active_user=True).first()
    if not user:
        return jsonify({'error': 'Nieznany kod pracownika.'}), 404

    carton = ImportedCarton.query.filter_by(barcode=package_barcode).first()
    if not carton:
        return jsonify({'error': 'Nieznany kod paczki.'}), 404

    if not carton.scan_start_at:
        return jsonify({'error': 'Brak zarejestrowanego startu dla tej paczki.'}), 400

    if carton.scan_end_at:
        return jsonify({'error': 'Paczka jest już zakończona.'}), 409

    # Tylko pracownik, który rozpoczął paczkę, może ją zakończyć
    if carton.scan_start_by and carton.scan_start_by != user.id:
        starter = User.query.get(carton.scan_start_by)
        return jsonify({
            'error': f'Paczkę rozpoczął {starter.display_name if starter else "inny pracownik"} — tylko on może ją zakończyć.'
        }), 403

    carton.scan_end_at = datetime.utcnow()
    carton.scan_end_by = user.id
    db.session.commit()

    secs = carton.processing_seconds()
    mins, s = divmod(secs, 60)
    time_str = f'{mins}m {s}s'

    return jsonify({
        'message': f'Koniec zarejestrowany — czas procesowania: {time_str}.',
        'carton': carton.to_dict()
    }), 200


# ══════════════════════════════════════════════════════════════════════════════
#  TIME TRACKING
# ══════════════════════════════════════════════════════════════════════════════

def _compute_worker_times(uid, shift, attendance_time):
    """Return break/work summary for one worker-shift pair."""
    events = WorkerTimeEvent.query.filter_by(
        user_id=uid, shift_id=shift.id
    ).order_by(WorkerTimeEvent.timestamp).all()

    break_secs = 0
    open_break = None
    work_end_ts = None
    breaks = []

    for e in events:
        if e.event_type == 'break_start':
            open_break = e.timestamp
        elif e.event_type == 'break_end' and open_break:
            secs = (e.timestamp - open_break).total_seconds()
            breaks.append({'start': iso_z(open_break), 'end': iso_z(e.timestamp),
                           'minutes': int(secs / 60)})
            break_secs += secs
            open_break = None
        elif e.event_type == 'work_end':
            work_end_ts = e.timestamp

    if open_break:
        breaks.append({'start': iso_z(open_break), 'end': None, 'minutes': None})

    end_ref = work_end_ts or datetime.utcnow()
    work_secs = max(0, (end_ref - attendance_time).total_seconds() - break_secs)

    return {
        'work_end':      iso_z(work_end_ts),
        'break_minutes': int(break_secs / 60),
        'work_minutes':  int(work_secs / 60),
        'breaks':        breaks,
        'on_break':      open_break is not None,
        'work_ended':    work_end_ts is not None,
        'events':        [e.to_dict() for e in events],
    }


@app.route('/time-tracking')
@leader_required
def time_tracking():
    return render_template('time_tracking.html')


@app.route('/worker-times')
@leader_required
def worker_times():
    return render_template('worker_times.html',
                           break_threshold=get_setting_int('break_threshold_minutes', 30))


@app.route('/api/time/scan', methods=['POST'])
@leader_required
def api_time_scan():
    data = json_body()
    barcode = (data.get('barcode') or '').strip()
    mode    = data.get('mode', 'break')   # 'break' | 'work_end'

    if not barcode:
        return jsonify({'error': 'Brak kodu.'}), 400

    user = User.query.filter_by(barcode_id=barcode, is_active_user=True).first()
    if not user:
        return jsonify({'error': 'Nieznany kod pracownika.'}), 404

    # Najnowsza obecność dziś
    today = local_today()
    attendance = ShiftAttendance.query.join(Shift).filter(
        ShiftAttendance.user_id == user.id,
        Shift.date == today
    ).order_by(ShiftAttendance.scanned_at.desc()).first()

    if not attendance:
        return jsonify({'error': f'{user.display_name} nie jest zeskanowany/a na zmianę dziś.'}), 400

    shift = attendance.shift
    events = WorkerTimeEvent.query.filter_by(
        user_id=user.id, shift_id=shift.id
    ).order_by(WorkerTimeEvent.timestamp).all()

    work_ended = any(e.event_type == 'work_end' for e in events)
    now = datetime.utcnow()

    if mode == 'work_end':
        if work_ended:
            return jsonify({'error': f'{user.display_name} już zakończył/a pracę na tej zmianie.'}), 409

        break_starts = sum(1 for e in events if e.event_type == 'break_start')
        break_ends   = sum(1 for e in events if e.event_type == 'break_end')
        if break_starts > break_ends:
            db.session.add(WorkerTimeEvent(
                user_id=user.id, shift_id=shift.id, event_type='break_end',
                timestamp=now, recorded_by=current_user.id, is_manual=False,
                note='Auto-zamknięcie przerwy przy końcu pracy'
            ))

        db.session.add(WorkerTimeEvent(
            user_id=user.id, shift_id=shift.id, event_type='work_end',
            timestamp=now, recorded_by=None, is_manual=False
        ))
        db.session.commit()
        return jsonify({'message': f'{user.display_name} — koniec pracy zarejestrowany.',
                        'event_type': 'work_end', 'user': user.to_dict()}), 200

    else:  # break
        if work_ended:
            return jsonify({'error': f'{user.display_name} już zakończył/a pracę.'}), 409

        break_starts = sum(1 for e in events if e.event_type == 'break_start')
        break_ends   = sum(1 for e in events if e.event_type == 'break_end')
        on_break     = break_starts > break_ends
        event_type   = 'break_end' if on_break else 'break_start'

        db.session.add(WorkerTimeEvent(
            user_id=user.id, shift_id=shift.id, event_type=event_type,
            timestamp=now, recorded_by=None, is_manual=False
        ))
        db.session.commit()

        label = 'wrócił/wróciła z przerwy ✅' if event_type == 'break_end' else 'poszedł/poszła na przerwę ☕'
        return jsonify({'message': f'{user.display_name} — {label}',
                        'event_type': event_type, 'user': user.to_dict()}), 200


@app.route('/api/worker-times')
@leader_required
def api_worker_times():
    date_str = request.args.get('date', local_today().isoformat())
    try:
        query_date = date.fromisoformat(date_str)
    except ValueError:
        query_date = local_today()

    shifts = Shift.query.filter_by(date=query_date).all()
    if not shifts:
        return jsonify({'workers': [], 'date': date_str}), 200

    shift_ids = [s.id for s in shifts]
    attendances = ShiftAttendance.query.filter(
        ShiftAttendance.shift_id.in_(shift_ids)
    ).options(joinedload(ShiftAttendance.user),
              joinedload(ShiftAttendance.shift)).all()

    # One entry per user — earliest attendance_in
    by_user = {}
    for att in attendances:
        uid = att.user_id
        if uid not in by_user or att.scanned_at < by_user[uid].scanned_at:
            by_user[uid] = att

    result = []
    for uid, att in by_user.items():
        summary = _compute_worker_times(uid, att.shift, att.scanned_at)
        summary.update({
            'user_id':      uid,
            'user_name':    att.user.display_name,
            'shift_id':     att.shift_id,
            'shift_number': att.shift.shift_number,
            'shift_in':     iso_z(att.scanned_at),
        })
        result.append(summary)

    result.sort(key=lambda x: x['user_name'])
    return jsonify({'workers': result, 'date': date_str}), 200


@app.route('/api/worker-times/event', methods=['POST'])
@leader_required
def api_time_event_create():
    data = json_body()
    user_id    = data.get('user_id')
    shift_id   = data.get('shift_id')
    event_type = data.get('event_type')
    ts_str     = data.get('timestamp', '')
    note       = (data.get('note') or '').strip()

    if not all([user_id, shift_id, event_type, ts_str]):
        return jsonify({'error': 'Brakuje wymaganych pól.'}), 400
    if event_type not in EVENT_TYPES:
        return jsonify({'error': 'Nieprawidłowy typ zdarzenia.'}), 400
    try:
        ts = datetime.fromisoformat(str(ts_str))
    except (TypeError, ValueError):
        return jsonify({'error': 'Nieprawidłowy format czasu.'}), 400

    # Bez tego bledne id konczylo sie naruszeniem klucza obcego i 500-ka
    # (na Postgresie transakcja padala w commicie).
    user_id = require_int(user_id, 'user_id')
    shift_id = require_int(shift_id, 'shift_id')
    if db.session.get(User, user_id) is None:
        return jsonify({'error': f'Nieznany pracownik (id {user_id}).'}), 400
    if db.session.get(Shift, shift_id) is None:
        return jsonify({'error': f'Nieznana zmiana (id {shift_id}).'}), 400

    event = WorkerTimeEvent(
        user_id=user_id, shift_id=shift_id, event_type=event_type,
        timestamp=ts, recorded_by=current_user.id, is_manual=True, note=note
    )
    db.session.add(event)
    db.session.commit()
    return jsonify(event.to_dict()), 201


@app.route('/api/worker-times/event/<int:event_id>', methods=['PUT'])
@leader_required
def api_time_event_update(event_id):
    event = WorkerTimeEvent.query.get_or_404(event_id)
    data  = json_body()
    if 'event_type' in data:
        if data['event_type'] not in EVENT_TYPES:
            return jsonify({'error': 'Nieprawidłowy typ.'}), 400
        event.event_type = data['event_type']
    if 'timestamp' in data:
        try:
            event.timestamp = datetime.fromisoformat(data['timestamp'])
        except ValueError:
            return jsonify({'error': 'Nieprawidłowy format czasu.'}), 400
    if 'note' in data:
        event.note = data['note']
    event.recorded_by = current_user.id
    event.is_manual   = True
    db.session.commit()
    return jsonify(event.to_dict()), 200


@app.route('/api/worker-times/event/<int:event_id>', methods=['DELETE'])
@leader_required
def api_time_event_delete(event_id):
    event = WorkerTimeEvent.query.get_or_404(event_id)
    db.session.delete(event)
    db.session.commit()
    return jsonify({'message': 'Usunięto.'}), 200


# ══════════════════════════════════════════════════════════════════════════════
#  FORECAST
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/forecast')
@leader_required
def forecast_page():
    today = local_today()
    date_from = (today - timedelta(days=7)).isoformat()
    date_to = (today + timedelta(days=14)).isoformat()
    return render_template('forecast.html', date_from=date_from, date_to=date_to)


@app.route('/api/forecast/chart-data')
@leader_required
def api_forecast_chart_data():
    date_from_str = request.args.get('date_from', '')
    date_to_str = request.args.get('date_to', '')

    try:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        date_from = local_today() - timedelta(days=7)

    try:
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        date_to = local_today() + timedelta(days=14)

    actual_rows = db.session.query(
        ImportedCarton.ziel_datum,
        func.sum(ImportedCarton.stueckzahl).label('total')
    ).filter(
        ImportedCarton.ziel_datum >= date_from,
        ImportedCarton.ziel_datum <= date_to,
        ImportedCarton.ziel_datum.isnot(None)
    ).group_by(ImportedCarton.ziel_datum).all()

    actual_map = {row.ziel_datum: (row.total or 0) for row in actual_rows}

    forecast_rows = Forecast.query.filter(
        Forecast.date >= date_from,
        Forecast.date <= date_to
    ).all()

    forecast_map = {f.date: {'quantity': f.quantity or 0, 'notes': f.notes or ''} for f in forecast_rows}

    result = []
    current = date_from
    while current <= date_to:
        forecast_qty = forecast_map.get(current, {}).get('quantity', 0)
        actual_qty = actual_map.get(current, 0)
        result.append({
            'date': current.isoformat(),
            'forecast': forecast_qty,
            'actual': actual_qty,
            'diff': forecast_qty - actual_qty,
            'notes': forecast_map.get(current, {}).get('notes', '')
        })
        current += timedelta(days=1)

    return jsonify(result)


@app.route('/api/forecast/save', methods=['POST'])
@leader_required
def api_forecast_save():
    data = request.get_json(silent=True)
    entries = data if isinstance(data, list) else [data or {}]

    saved = 0
    for entry in entries:
        if not isinstance(entry, dict):
            continue
        date_str = str(entry.get('date') or '').strip()
        try:
            d = datetime.strptime(date_str, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            continue

        qty = require_int(entry.get('quantity') or 0, 'quantity')
        notes = (entry.get('notes') or '').strip()

        existing = Forecast.query.filter_by(date=d).first()
        if existing:
            existing.quantity = qty
            existing.notes = notes
            existing.updated_at = datetime.utcnow()
            existing.updated_by = current_user.id
        else:
            db.session.add(Forecast(
                date=d,
                quantity=qty,
                notes=notes,
                created_by=current_user.id
            ))
        saved += 1

    db.session.commit()
    return jsonify({'message': f'Zapisano {saved} wpisów.', 'saved': saved})


@app.route('/api/forecast/export')
@leader_required
def api_forecast_export():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.chart import BarChart, Reference
    from flask import make_response
    import io as _io

    date_from_str = request.args.get('date_from', '')
    date_to_str = request.args.get('date_to', '')

    try:
        date_from = datetime.strptime(date_from_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        date_from = local_today() - timedelta(days=7)

    try:
        date_to = datetime.strptime(date_to_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        date_to = local_today() + timedelta(days=14)

    actual_rows = db.session.query(
        ImportedCarton.ziel_datum,
        func.sum(ImportedCarton.stueckzahl).label('total')
    ).filter(
        ImportedCarton.ziel_datum >= date_from,
        ImportedCarton.ziel_datum <= date_to,
        ImportedCarton.ziel_datum.isnot(None)
    ).group_by(ImportedCarton.ziel_datum).all()

    actual_map = {row.ziel_datum: (row.total or 0) for row in actual_rows}

    forecast_rows = Forecast.query.filter(
        Forecast.date >= date_from,
        Forecast.date <= date_to
    ).all()

    forecast_map = {f.date: {'quantity': f.quantity or 0, 'notes': f.notes or ''} for f in forecast_rows}

    days = []
    current = date_from
    while current <= date_to:
        days.append(current)
        current += timedelta(days=1)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Forecast vs Actual'

    header_font = Font(bold=True, color='FFFFFF')
    col_fills = [
        PatternFill('solid', fgColor='2D3250'),
        PatternFill('solid', fgColor='4A6FD9'),
        PatternFill('solid', fgColor='27AE60'),
        PatternFill('solid', fgColor='C0392B'),
        PatternFill('solid', fgColor='7D3C98'),
    ]
    headers = ['Data', 'Forecast', 'Actual (Paczki)', 'Różnica (F-A)', 'Notatki']

    for col, (h, fill) in enumerate(zip(headers, col_fills), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center')

    for row_idx, d in enumerate(days, start=2):
        forecast_qty = forecast_map.get(d, {}).get('quantity', 0)
        actual_qty = actual_map.get(d, 0)
        diff = forecast_qty - actual_qty

        ws.cell(row=row_idx, column=1, value=d.strftime('%d.%m.%Y'))
        ws.cell(row=row_idx, column=2, value=forecast_qty)
        ws.cell(row=row_idx, column=3, value=actual_qty)
        diff_cell = ws.cell(row=row_idx, column=4, value=diff)
        ws.cell(row=row_idx, column=5, value=forecast_map.get(d, {}).get('notes', ''))

        if diff > 0:
            diff_cell.fill = PatternFill('solid', fgColor='FDEBD0')
        elif diff < 0:
            diff_cell.fill = PatternFill('solid', fgColor='FADBD8')

    for col_letter, width in [('A', 14), ('B', 14), ('C', 18), ('D', 18), ('E', 35)]:
        ws.column_dimensions[col_letter].width = width

    n = len(days)
    if n > 0:
        chart = BarChart()
        chart.type = 'col'
        chart.grouping = 'clustered'
        chart.title = 'Forecast vs Actual'
        chart.y_axis.title = 'Ilość paczek'
        chart.x_axis.title = 'Data'
        chart.width = 26
        chart.height = 14

        forecast_ref = Reference(ws, min_col=2, min_row=1, max_row=n + 1)
        actual_ref = Reference(ws, min_col=3, min_row=1, max_row=n + 1)
        dates_ref = Reference(ws, min_col=1, min_row=2, max_row=n + 1)

        chart.add_data(forecast_ref, titles_from_data=True)
        chart.add_data(actual_ref, titles_from_data=True)
        chart.set_categories(dates_ref)
        chart.series[0].graphicalProperties.solidFill = '4A6FD9'
        chart.series[1].graphicalProperties.solidFill = '27AE60'

        ws.add_chart(chart, 'G2')

    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = make_response(buf.read())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    resp.headers['Content-Disposition'] = (
        f'attachment; filename=forecast_{date_from_str}_{date_to_str}.xlsx'
    )
    return resp


# ══════════════════════════════════════════════════════════════════════════════
#  SEED DATA
# ══════════════════════════════════════════════════════════════════════════════

DEFAULT_ACTIVITIES = [
    'Post Processing',
    'Zwroty',
    'Organic Decoration',
    'Rollout Decoration',
    'Expansion',
    'Textile-Picking',
    'Order-VAS',
    'Carton Labeling',
    'Orders',
]


DEFAULT_COUNTRY_MAPPINGS = [
    ('Schweiz', '91000741810'),
    ('Italien', '91000741812'),
    ('Rumänien', '91000741814'),
    ('Vereinigtes Königreich', '91000741816'),
    ('Frankreich', '91000741817'),
    ('Croatia', '91000741820'),
    ('Spanien Kanaren', 'ES01'),
    ('Spanien', 'ES01'),
    ('Portugal', '91000741824'),
    ('Elfenbeinküste', '91000741828'),
    ('Deutschland', 'Orsay DE'),
    ('Kongo', 'IAM RCB'),
    ('Senegal', 'IAM SN'),
    ('Deutschland AMAZON', 'DE AMAZON'),
    ('EDEKA', 'EDK1'),
    ('Netherlands', 'NL01'),
    ('Northern Ireland', 'IRL'),
    ('ES', 'ES01'),
    ('IT', '91000741812'),
    ('CH', '91000741810'),
    ('Slowakei', 'SLO'),
    ('Tschechien', 'TSC'),
    ('PL', 'PL'),
    ('HU', 'HU'),
    ('BE', 'BE'),
    ('PT', 'PT'),
    ('AT', 'AT'),
    ('Deutschland C&A', 'DE'),
    ('SI', 'SI'),
]


def seed_data():
    """Seed default activities, admin user and country mappings on first run."""
    if Activity.query.count() == 0:
        for i, name in enumerate(DEFAULT_ACTIVITIES):
            db.session.add(Activity(name=name, sort_order=i))
        db.session.commit()
        print("[SEED] Default activities created.")

    if User.query.filter_by(role='admin').count() == 0:
        haslo = os.environ.get('ADMIN_PASSWORD', '').strip() or DEFAULT_ADMIN_PASSWORD
        admin = User(
            username='admin',
            display_name='Administrator',
            role='admin'
        )
        admin.set_password(haslo)
        db.session.add(admin)
        db.session.commit()
        if haslo == DEFAULT_ADMIN_PASSWORD:
            print("[SEED] Admin user created (login: admin / password: "
                  f"{DEFAULT_ADMIN_PASSWORD})")
            print("[SEED] !!! ZMIEN TO HASLO PRZY PIERWSZYM LOGOWANIU (/profile) "
                  "albo ustaw ADMIN_PASSWORD przed pierwszym startem.")
        else:
            print("[SEED] Admin user created (login: admin / haslo z ADMIN_PASSWORD)")

    if CountryMapping.query.count() == 0:
        for country, innenauftrag in DEFAULT_COUNTRY_MAPPINGS:
            db.session.add(CountryMapping(country=country, innenauftrag=innenauftrag))
        db.session.commit()
        print("[SEED] Default country mappings created.")


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def migrate_columns():
    """Add missing columns to existing tables without dropping data.

    The ALTER TABLE block is SQLite-only: it patches databases created before a
    column existed. On Postgres the schema always comes from db.create_all(), so
    there is nothing to patch — and a failed statement there aborts the whole
    transaction, hence the rollback in the except branches. Indexes are created
    on both engines (CREATE INDEX IF NOT EXISTS works on either).
    """
    is_sqlite = db.engine.dialect.name == 'sqlite'
    with db.engine.connect() as conn:
        migrations = [] if not is_sqlite else [
            ("imported_carton", "processed_by",   "INTEGER REFERENCES user(id)"),
            ("imported_carton", "processed_at",   "DATETIME"),
            ("general_stat",    "double_rate_category_data", "TEXT DEFAULT '{}'"),
            ("imported_carton", "double_rate",     "BOOLEAN DEFAULT 0"),
            ("imported_carton", "scan_start_at",  "DATETIME"),
            ("imported_carton", "scan_start_by",  "INTEGER REFERENCES user(id)"),
            ("imported_carton", "scan_end_at",    "DATETIME"),
            ("imported_carton", "scan_end_by",    "INTEGER REFERENCES user(id)"),
            ("imported_carton", "added_manually", "BOOLEAN DEFAULT 0"),
            ("imported_carton", "modified_at",    "DATETIME"),
            ("imported_carton", "modified_by",    "INTEGER REFERENCES user(id)"),
        ]
        for table, column, col_def in migrations:
            try:
                conn.execute(db.text(f"ALTER TABLE {table} ADD COLUMN {column} {col_def}"))
                conn.commit()
            except Exception:
                conn.rollback()  # column already exists

        indexes = [
            "CREATE INDEX IF NOT EXISTS ix_carton_ziel_datum   ON imported_carton (ziel_datum)",
            "CREATE INDEX IF NOT EXISTS ix_carton_uebergabe_nr ON imported_carton (uebergabe_nr)",
            "CREATE INDEX IF NOT EXISTS ix_carton_processed_by ON imported_carton (processed_by)",
            "CREATE INDEX IF NOT EXISTS ix_carton_land         ON imported_carton (land)",
            "CREATE INDEX IF NOT EXISTS ix_carton_imported_at  ON imported_carton (imported_at)",
            "CREATE INDEX IF NOT EXISTS ix_gstat_loading_date  ON general_stat    (loading_date)",
            "CREATE INDEX IF NOT EXISTS ix_gstat_list_id       ON general_stat    (list_id)",
            "CREATE INDEX IF NOT EXISTS ix_wte_user_shift      ON worker_time_event (user_id, shift_id)",
            "CREATE INDEX IF NOT EXISTS ix_wte_shift_id        ON worker_time_event (shift_id)",
        ]
        for sql in indexes:
            try:
                conn.execute(db.text(sql))
                conn.commit()
            except Exception:
                conn.rollback()


def init_db():
    """Create the schema, patch columns and seed — safe to run from every worker.

    Gunicorn imports this module once per worker, so on a fresh Postgres all
    workers race inside create_all() and the loser dies with a UniqueViolation on
    pg_class ("worker failed to boot"). A Postgres advisory lock serializes them:
    the winner does the DDL, the rest wait and then find nothing left to do
    (create_all and seed_data are both no-ops on an initialized database).
    SQLite needs no lock — the writer lock plus busy_timeout already serialize it.
    """
    if db.engine.dialect.name != 'postgresql':
        db.create_all()
        migrate_columns()
        seed_data()
        return

    lock_id = 5001  # dowolna stala — byle ta sama we wszystkich workerach
    with db.engine.connect() as conn:
        conn.execute(db.text('SELECT pg_advisory_lock(:id)'), {'id': lock_id})
        conn.commit()
        try:
            db.create_all()
            migrate_columns()
            seed_data()
        finally:
            conn.execute(db.text('SELECT pg_advisory_unlock(:id)'), {'id': lock_id})
            conn.commit()


with app.app_context():
    init_db()

if __name__ == '__main__':
    # Debugger Werkzeuga pozwala wykonac dowolny kod, wiec domyslnie sluchamy
    # tylko na loopbacku. Dostep z LAN-u w dev: HOST=0.0.0.0 python app.py
    # (produkcja i tak chodzi na gunicornie z Dockerfile, nie tedy).
    app.run(
        debug=os.environ.get('FLASK_DEBUG', '1') == '1',
        host=os.environ.get('HOST', '127.0.0.1'),
        port=int(os.environ.get('PORT', 5001)),
    )
